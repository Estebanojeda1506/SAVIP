"""Pruebas del cálculo de empalme ICCP -> ICOCIV."""

from __future__ import annotations

import os
import sys
from pathlib import Path


os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app_icociv.servicios.empalme_iccp_icociv import (  # noqa: E402
    calcular_empalme_acero,
    calcular_empalme_general,
    series_iccp_por_tipo,
)


ICCP = {
    "Grupo": {"2021_1": 90, "2021_6": 95, "2021_12": 100},
    "Total ICCP": {"2021_1": 100, "2021_6": 105, "2021_12": 110},
    "Equipos": {"2021_1": 100, "2021_6": 105, "2021_12": 110},
}
ICOCIV = {"2021_12": 100, "2022_1": 120, "2022_6": 150}


def test_empalme_completo_general() -> None:
    resultado = calcular_empalme_general(
        {
            "precio_base": 100,
            "anticipo_amortizado": 10,
            "fecha_inicial": "2021-01",
            "fecha_final": "2022-01",
            "grupo_iccp": "Grupo",
            "ruta_icociv": "Ruta",
        },
        ICOCIV,
        ICCP,
    )

    assert round(resultado["r1"], 6) == 10
    assert round(resultado["r2"], 6) == 20
    assert round(resultado["r_total"], 6) == round(resultado["r1"] + resultado["r2"], 6)
    assert round(resultado["valor_actualizado"], 6) == 120


def test_solo_iccp_general() -> None:
    resultado = calcular_empalme_general(
        {
            "precio_base": 100,
            "fecha_inicial": "2021-01",
            "fecha_final": "2021-06",
            "grupo_iccp": "Grupo",
        },
        {},
        ICCP,
    )

    assert round(resultado["r1"], 6) == round(100 * ((95 / 90) - 1), 6)
    assert resultado["r2"] == 0


def test_iccp_canasta_general_activa() -> None:
    resultado = calcular_empalme_general(
        {
            "precio_base": 100,
            "fecha_inicial": "2021-01",
            "fecha_final": "2021-06",
            "tipo_serie_iccp": "canasta_general",
            "serie_iccp": "Equipos",
        },
        {},
        ICCP,
    )

    assert resultado["tipo_serie_iccp"] == "canasta_general"
    assert resultado["serie_iccp"] == "Equipos"
    assert round(resultado["r1"], 6) == 5


def test_iccp_separa_total_canasta_y_grupo() -> None:
    series = series_iccp_por_tipo(ICCP)
    assert series["total_iccp"] == ["Total ICCP"]
    assert series["canasta_general"] == ["Equipos"]
    assert series["grupo_obra"] == ["Grupo"]


def test_iccp_rechaza_serie_fuera_de_tipo() -> None:
    try:
        calcular_empalme_general(
            {
                "precio_base": 100,
                "fecha_inicial": "2021-01",
                "fecha_final": "2021-06",
                "tipo_serie_iccp": "canasta_general",
                "serie_iccp": "Total ICCP",
            },
            {},
            ICCP,
        )
    except ValueError as exc:
        assert "no pertenece" in str(exc)
    else:
        raise AssertionError("Debía rechazar Total ICCP dentro de Canasta general.")


def test_iccp_rechaza_canasta_y_grupo_simultaneos() -> None:
    try:
        calcular_empalme_general(
            {
                "precio_base": 100,
                "fecha_inicial": "2021-01",
                "fecha_final": "2021-06",
                "canasta_general_iccp": "Total ICCP",
                "grupo_obra_iccp": "Grupo",
            },
            {},
            ICCP,
        )
    except ValueError as exc:
        assert "solo una serie ICCP" in str(exc)
    else:
        raise AssertionError("Debía rechazar dos series ICCP activas.")


def test_acero_calcula_z() -> None:
    resultado = calcular_empalme_acero(
        {
            "p0": 100,
            "ix": 2,
            "q": 70,
            "fecha_inicial": "2021-01",
            "fecha_final": "2022-01",
            "grupo_iccp": "Grupo",
            "ruta_icociv": "Ruta",
            "calculo_acero": True,
        },
        ICOCIV,
        ICCP,
    )

    assert round(resultado["r_total"], 6) == round(100 * ((100 / 90) - 1) + (100 + 100 * ((100 / 90) - 1)) * 0.2, 6)
    assert round(resultado["r_total"], 6) == round(resultado["r1"] + resultado["r2"], 6)
    assert round(resultado["z"], 6) == round(140 - (resultado["r_total"] + 100), 6)


def test_acero_calcula_r_sin_ix_q_y_deja_z_pendiente() -> None:
    resultado = calcular_empalme_acero(
        {
            "p0": 100,
            "ix": 0,
            "q": 0,
            "fecha_inicial": "2021-01",
            "fecha_final": "2022-01",
            "grupo_iccp": "Grupo",
            "ruta_icociv": "Ruta",
            "calculo_acero": True,
        },
        ICOCIV,
        ICCP,
    )

    assert round(resultado["r_total"], 6) == round(resultado["r1"] + resultado["r2"], 6)
    assert resultado["valor_facturado_total"] is None
    assert resultado["z"] is None
    assert "Ix y q" in resultado["z_observacion"]


def test_exportacion_excel_empalme_hoja_unica_y_formulas(tmp_path) -> None:
    from openpyxl import load_workbook
    from PySide6.QtWidgets import QApplication

    from app_icociv.interfaz.widgets.empalme_iccp_icociv import WidgetEmpalmeICCPICOCIV

    _app = QApplication.instance() or QApplication([])
    widget = WidgetEmpalmeICCPICOCIV()
    iccp = {
        "Materiales": {"2021_1": 100, "2021_12": 110},
        "Total ICCP": {"2021_1": 100, "2021_12": 110},
    }
    icociv = {"2021_12": 100, "2022_1": 120, "2022_2": 125}
    widget.calculos = [
        calcular_empalme_general(
            {
                "item": "Cemento",
                "unidad": "m3",
                "precio_base": 1000,
                "fecha_inicial": "2021-01",
                "fecha_final": "2022-01",
                "tipo_serie_iccp": "canasta_general",
                "serie_iccp": "Materiales",
                "ruta_icociv": "Ruta > Cemento",
            },
            icociv,
            iccp,
        ),
        calcular_empalme_general(
            {
                "item": "Arena",
                "unidad": "kg",
                "precio_base": 500,
                "fecha_inicial": "2021-01",
                "fecha_final": "2022-02",
                "tipo_serie_iccp": "total_iccp",
                "serie_iccp": "Total ICCP",
                "ruta_icociv": "Ruta > Arena",
            },
            icociv,
            iccp,
        ),
    ]
    widget.calculos.append(
        calcular_empalme_acero(
            {
                "item": "Acero de refuerzo",
                "unidad": "kg",
                "p0": 800,
                "ix": 5.5,
                "q": 200,
                "fecha_inicial": "2021-01",
                "fecha_final": "2022-01",
                "tipo_serie_iccp": "total_iccp",
                "serie_iccp": "Total ICCP",
                "ruta_icociv": "Ruta > Acero",
                "calculo_acero": True,
            },
            icociv,
            iccp,
        )
    )
    ruta = tmp_path / "empalme.xlsx"
    widget._generar_excel_empalme(str(ruta))

    wb = load_workbook(ruta, data_only=False)
    assert wb.sheetnames == ["Empalme ICCP-ICOCIV"]
    ws = wb["Empalme ICCP-ICOCIV"]
    encabezado = next(
        fila
        for fila in range(1, ws.max_row + 1)
        if ws.cell(fila, 1).value == "INSUMO CONTRACTUAL" and ws.cell(fila, 2).value == "UNIDAD"
    )
    for fila in (encabezado + 1, encabezado + 2, encabezado + 3):
        assert isinstance(ws.cell(fila, 3).value, (int, float))
        assert isinstance(ws.cell(fila, 4).value, (int, float))
        assert isinstance(ws.cell(fila, 5).value, (int, float))
        assert ws.cell(fila, 6).value == f'=IF(OR(D{fila}="",E{fila}=""),0,(C{fila}-L{fila})*((E{fila}/D{fila})-1))'
        assert ws.cell(fila, 9).value == f'=IF(OR(G{fila}="",H{fila}=""),0,((C{fila}-L{fila})+F{fila})*((H{fila}/G{fila})-1))'
        assert ws.cell(fila, 10).value == f"=F{fila}+I{fila}"
        assert ws.cell(fila, 11).value == f"=(C{fila}-L{fila})+J{fila}"

    # El detalle de acero referencia la fila del mismo cálculo dentro de la misma hoja.
    fila_acero_encabezado = next(
        fila
        for fila in range(encabezado + 1, ws.max_row + 1)
        if ws.cell(fila, 1).value == "INSUMO CONTRACTUAL" and ws.cell(fila, 2).value == "P0"
    )
    fila_acero = fila_acero_encabezado + 1
    fila_calculo_acero = encabezado + 3
    assert ws.cell(fila_acero, 6).value == f"=F{fila_calculo_acero}"
    assert ws.cell(fila_acero, 7).value == f"=I{fila_calculo_acero}"
    assert ws.cell(fila_acero, 8).value == f"=J{fila_calculo_acero}"
    assert ws.cell(fila_acero, 10).value == f'=IF(E{fila_acero}="","No calculado",E{fila_acero}-(H{fila_acero}+B{fila_acero}))'


if __name__ == "__main__":
    from tempfile import TemporaryDirectory

    test_empalme_completo_general()
    test_solo_iccp_general()
    test_iccp_canasta_general_activa()
    test_iccp_separa_total_canasta_y_grupo()
    test_iccp_rechaza_serie_fuera_de_tipo()
    test_iccp_rechaza_canasta_y_grupo_simultaneos()
    test_acero_calcula_z()
    test_acero_calcula_r_sin_ix_q_y_deja_z_pendiente()
    with TemporaryDirectory() as carpeta:
        test_exportacion_excel_empalme_hoja_unica_y_formulas(Path(carpeta))
    print("OK: empalme ICCP -> ICOCIV")
