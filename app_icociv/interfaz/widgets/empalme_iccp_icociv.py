"""Widget independiente para actualización de precios ICCP -> ICOCIV."""

from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any, Callable
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import Qt
from PySide6.QtWidgets import (
    QAbstractSpinBox,
    QCheckBox,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSpinBox,
    QTableView,
    QTextBrowser,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from app_icociv.interfaz.controladores.controlador_principal import ControladorPrincipal
from app_icociv.interfaz.estilos.constantes_visuales import ALTURA_CONTROL, ANCHO_BOTON_INCREMENTO
from app_icociv.interfaz.widgets.modelo_tabla import ModeloTablaPandas
from app_icociv.proyeccion.servicio_proyeccion import resolver_fila_seleccionada
from app_icociv.servicios.empalme_iccp_icociv import calcular_empalme_iccp_icociv, normalizar_periodo_empalme, series_iccp_por_tipo
from app_icociv.utilidades.nomenclatura_icociv import nombre_nivel, texto_checkbox


AYUDAS = {
    "fecha_inicial": "Fecha desde la cual se inicia la actualización del precio. Puede corresponder a fecha base contractual, cierre del proceso licitatorio, propuesta, firma, no objeción u otra fecha técnicamente justificada.",
    "fecha_final": "Fecha hasta la cual se desea actualizar el precio. Puede corresponder al mes de ejecución, acta, comparación o mes final definido por el usuario.",
    "precio": "Valor inicial que se desea actualizar. Puede ser el precio de un insumo, ítem, APU, acta parcial o componente contractual.",
    "anticipo": "Valor del anticipo ya amortizado que debe descontarse de la base de cálculo. Si no aplica, ingrese cero.",
    "base": "Base ajustable = P - A. Es el valor realmente sujeto a actualización.",
    "cantidad": "Cantidad del insumo o ítem actualizado. Puede usarse para calcular valores totales cuando aplique.",
    "unidad": "Unidad de medida: KG, M3, M2, ML, unidad, global, entre otras.",
    "grupo_iccp": "Grupo del ICCP usado para el primer tramo. Debe seleccionarse según criterio técnico del ingeniero.",
    "ruta_icociv": "Ruta jerárquica ICOCIV equivalente al grupo ICCP o insumo analizado. La selección es manual y debe justificarse técnicamente.",
    "observacion": "Justifique la equivalencia, fecha usada o cualquier criterio técnico aplicado en el empalme.",
    "p0": "P0 corresponde al valor base del insumo acero que será ajustado: ofertado, contractual o valor de acta parcial, según el caso.",
    "ix": "Ix corresponde al valor real facturado por kilogramo de acero, según factura, cotización u orden de compra validada.",
    "q": "q corresponde a la cantidad de kilogramos de acero ejecutados o reconocidos en el periodo analizado.",
}

AYUDAS.update(
    {
        "tipo_serie_iccp": "Define si el cálculo ICCP usará Total ICCP, Canasta general o Grupo de obra. Solo puede usarse una categoría.",
        "serie_iccp": "Serie ICCP activa tomada del Anexo 10 histórico; se usa para calcular I0 ICCP, I ICCP, factor ICCP y R1.",
        "canasta_general_iccp": "Serie general del ICCP tomada del Anexo 10 histórico.",
        "z": "Valor adicional por fluctuación del acero, cuando aplique.",
    }
)

COLUMNAS_EQUIVALENCIAS = [
    ("insumo_contractual", "INSUMO CONTRACTUAL"),
    ("grupo_iccp_equivalente", "GRUPO DE OBRA EQUIVALENTE ICCP"),
    ("insumo_icociv_equivalente", "INSUMO EQUIVALENTE ICOCIV"),
]

COLUMNAS_VALOR_AJUSTADO = [
    ("insumo_contractual", "INSUMO CONTRACTUAL"),
    ("unidad", "UNIDAD"),
    ("precio_base", "PRECIO INSUMO CONTRACTUAL-APU"),
    ("i0_iccp", "ÍNDICE ICCP INICIAL (I0)"),
    ("i_iccp", "ÍNDICE ICCP FINAL (I)"),
    ("r1", "R1"),
    ("i0_icociv", "ÍNDICE ICOCIV INICIAL (I0)"),
    ("i_icociv", "ÍNDICE ICOCIV FINAL (I)"),
    ("r2", "R2"),
    ("r_total", "VALOR DEL AJUSTE (R)"),
    ("valor_actualizado", "VALOR INSUMO AJUSTADO"),
]

COLUMNAS_DETALLE_EXPORTACION = [
    ("calculo_id", "ID cálculo"),
    ("numero_calculo", "N°"),
    ("codigo_item", "Código"),
    ("tipo_calculo", "Tipo"),
    ("fecha_inicial", "Fecha inicial"),
    ("fecha_final", "Fecha final"),
    ("ruta_iccp", "Ruta ICCP"),
    ("ruta_icociv", "Ruta ICOCIV"),
    ("diferencia_absoluta", "Diferencia absoluta"),
    ("icociv_final_es_proyectado", "ICOCIV final proyectado"),
    ("ultimo_periodo_icociv_real", "Último periodo ICOCIV real"),
    ("indice_icociv_final_proyectado", "Índice ICOCIV final proyectado"),
    ("modelo_proyeccion", "Modelo proyección"),
    ("horizonte_usado", "Horizonte usado"),
    ("estado_horizonte", "Estado horizonte"),
    ("advertencias_proyeccion", "Advertencias proyección"),
    ("observacion_tecnica", "Observación técnica"),
    ("fecha_calculo", "Fecha cálculo"),
    ("trazabilidad_formula", "Desarrollo matemático"),
]

UNIDADES_INGENIERIA = ["m lineal", "m2", "m3", "kg", "Unidad", "Global"]
HOJA_EXPORTACION_EMPALME = "Empalme ICCP-ICOCIV"
FORMATO_MONEDA_EXCEL = '$ #,##0.00'
FORMATO_INDICE_EXCEL = '#,##0.00'
FORMATO_FACTOR_EXCEL = '0.000000'


from app_icociv.interfaz.widgets.controles import (
    ComboBoxSinRueda,
    SpinEnteroSinRueda,
    SpinSinRueda,
)


class WidgetEmpalmeICCPICOCIV(QWidget):
    """Pestaña de cálculo contractual con trazabilidad acumulativa."""

    def __init__(self) -> None:
        super().__init__()
        self.controlador_icociv = ControladorPrincipal()
        self.calculos: list[dict[str, Any]] = []
        self.indice_edicion: int | None = None
        self.callback_proyeccion_icociv: Callable[[dict[str, Any], int, int], dict[str, Any]] | None = None
        self.modelo_equivalencias = ModeloTablaPandas()
        self.modelo_valor_ajustado = ModeloTablaPandas()
        self._crear_interfaz()
        self._actualizar_controles_selector(False)

    def configurar_proyeccion_icociv(self, callback: Callable[[dict[str, Any], int, int], dict[str, Any]]) -> None:
        """Conecta el empalme con el flujo de proyección ICOCIV existente."""
        self.callback_proyeccion_icociv = callback

    def actualizar_icociv(self, tablas: dict[str, pd.DataFrame], periodos: list[str]) -> None:
        """Carga el mismo árbol jerárquico ICOCIV usado por la app principal."""
        self.controlador_icociv.tablas = tablas
        self.controlador_icociv.periodos = periodos
        self.controlador_icociv.fuente_actual = None
        self._llenar_combo(self.combo_grupo_icociv, self.controlador_icociv.opciones_grupo())
        self._actualizar_controles_selector(True)
        self._actualizar_jerarquia_icociv()

    def _crear_interfaz(self) -> None:
        layout_principal = QVBoxLayout(self)
        layout_principal.setContentsMargins(0, 0, 0, 0)

        contenido = QWidget()
        layout = QVBoxLayout(contenido)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(12)

        titulo = QLabel("Actualización de precios / Empalme ICCP-ICOCIV")
        titulo.setObjectName("titulo_dashboard")
        subtitulo = QLabel(
            "Módulo independiente para contratos, ítems, insumos o APU. "
            "El ICCP histórico está cargado internamente desde el Anexo 10."
        )
        subtitulo.setWordWrap(True)
        layout.addWidget(titulo)
        layout.addWidget(subtitulo)

        layout.addWidget(self._grupo_datos_generales())
        layout.addWidget(self._grupo_periodo())
        layout.addWidget(self._grupo_calculo())
        layout.addWidget(self._grupo_selector_iccp())
        layout.addWidget(self._grupo_selector_icociv())
        layout.addWidget(self._grupo_observacion())
        layout.addWidget(self._grupo_acero())
        layout.addWidget(self._grupo_boton())
        layout.addWidget(self._grupo_resultado())
        layout.addWidget(self._grupo_tabla(), 1)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(contenido)
        layout_principal.addWidget(scroll)

    def _grupo_datos_generales(self) -> QGroupBox:
        grupo = QGroupBox("1. Datos generales del análisis")
        form = QFormLayout(grupo)
        self.contrato = QLineEdit()
        self.objeto = QTextEdit()
        self.objeto.setMaximumHeight(64)
        self.responsable = QLineEdit()
        self.observacion_general = QTextEdit()
        self.observacion_general.setMaximumHeight(64)
        _fila(form, "Nombre del contrato", self.contrato, "Identificación corta del contrato o análisis.")
        _fila(form, "Objeto / descripción", self.objeto, "Descripción breve del contrato, actividad o conjunto de ítems evaluados.")
        _fila(form, "Responsable técnico", self.responsable, "Nombre del profesional que prepara o revisa el cálculo.")
        _fila(form, "Observación general", self.observacion_general, "Notas generales del análisis, alcance o criterio contractual.")
        return grupo

    def _grupo_periodo(self) -> QGroupBox:
        grupo = QGroupBox("2. Periodo de análisis")
        form = QFormLayout(grupo)
        self.fecha_inicial_anio, self.fecha_inicial_mes = _periodo_spins(2021, 1)
        self.fecha_final_anio, self.fecha_final_mes = _periodo_spins(2024, 8)
        _fila(form, "Fecha inicial", _periodo_widget(self.fecha_inicial_anio, self.fecha_inicial_mes), AYUDAS["fecha_inicial"])
        _fila(form, "Fecha final", _periodo_widget(self.fecha_final_anio, self.fecha_final_mes), AYUDAS["fecha_final"])
        return grupo

    def _grupo_calculo(self) -> QGroupBox:
        grupo = QGroupBox("3. Datos del valor a actualizar")
        grid = QGridLayout(grupo)
        self.item = QLineEdit()
        self.codigo = QLineEdit()
        self.unidad = ComboBoxSinRueda()
        self._llenar_combo(self.unidad, [{"texto": unidad, "indice": unidad} for unidad in UNIDADES_INGENIERIA])
        self.cantidad = _spin(0, 1_000_000_000, 4)
        self.precio_base = _spin(0, 1_000_000_000_000_000, 2)
        self.anticipo = _spin(0, 1_000_000_000_000_000, 2)
        self.precio_base.lineEdit().setPlaceholderText("Ej.: 423467.54")
        self.anticipo.lineEdit().setPlaceholderText("Ej.: 0")
        campos = [
            ("Insumo / ítem", self.item, "Nombre del insumo, actividad, APU o componente contractual."),
            ("Código", self.codigo, "Código interno o contractual del ítem. Es opcional."),
            ("Unidad", self.unidad, AYUDAS["unidad"]),
            ("Cantidad", self.cantidad, AYUDAS["cantidad"]),
            ("Precio o valor base P", self.precio_base, AYUDAS["precio"]),
            ("Anticipo amortizado A", self.anticipo, AYUDAS["anticipo"]),
        ]
        for fila, (etiqueta, widget, ayuda) in enumerate(campos):
            grid.addWidget(_etiqueta_ayuda(etiqueta, ayuda), fila, 0)
            grid.addWidget(_ayuda(widget, ayuda), fila, 1)
        grid.setColumnStretch(1, 1)
        return grupo

    def _grupo_selector_iccp(self) -> QGroupBox:
        grupo = QGroupBox("4. Serie ICCP")
        form = QFormLayout(grupo)
        self.series_iccp = series_iccp_por_tipo()
        self.combo_tipo_iccp = ComboBoxSinRueda()
        self.combo_tipo_iccp.addItem("Sin selección", None)
        self.combo_tipo_iccp.addItem("Total ICCP", "total_iccp")
        self.combo_tipo_iccp.addItem("Canasta general", "canasta_general")
        self.combo_tipo_iccp.addItem("Grupo de obra", "grupo_obra")
        self.combo_tipo_iccp.currentIndexChanged.connect(lambda _: self._actualizar_series_iccp())
        self.combo_serie_iccp = ComboBoxSinRueda()
        self.combo_serie_iccp.setEnabled(False)
        _fila(form, "Tipo de serie ICCP", self.combo_tipo_iccp, AYUDAS["tipo_serie_iccp"])
        _fila(form, "Serie ICCP", self.combo_serie_iccp, AYUDAS["serie_iccp"])
        self._actualizar_series_iccp()
        return grupo

    def _grupo_observacion(self) -> QGroupBox:
        grupo = QGroupBox("6. Observación técnica del empalme")
        form = QFormLayout(grupo)
        self.observacion_item = QTextEdit()
        self.observacion_item.setMaximumHeight(70)
        _fila(form, "Observación técnica", self.observacion_item, AYUDAS["observacion"])
        return grupo

    def _grupo_acero(self) -> QGroupBox:
        grupo = QGroupBox("7. Cálculo especial para acero")
        form = QFormLayout(grupo)
        self.chk_acero = QCheckBox("Cálculo especial para acero")
        self.chk_acero.setToolTip("Active este cálculo cuando aplique la fórmula especial de fluctuación del acero.")
        self.chk_acero.toggled.connect(self._alternar_acero)
        self.p0 = _spin(0, 1_000_000_000_000_000, 2)
        self.ix = _spin(0, 1_000_000_000, 4)
        self.q = _spin(0, 1_000_000_000, 4)
        form.addRow(self.chk_acero)
        _fila(form, "P0 — Valor base del acero", self.p0, AYUDAS["p0"])
        _fila(form, "Ix — Valor facturado por kilogramo", self.ix, AYUDAS["ix"])
        _fila(form, "q — Cantidad ejecutada en kg", self.q, AYUDAS["q"])
        self._alternar_acero(False)
        return grupo

    def _grupo_boton(self) -> QWidget:
        contenedor = QWidget()
        layout = QHBoxLayout(contenedor)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.addStretch()
        self.boton_limpiar = QPushButton("Limpiar formulario")
        self.boton_limpiar.clicked.connect(self.limpiar_formulario)
        self.boton_calcular = QPushButton("Calcular actualización")
        self.boton_calcular.clicked.connect(self.calcular)
        layout.addWidget(self.boton_limpiar)
        layout.addWidget(self.boton_calcular)
        return contenedor

    def _grupo_selector_icociv(self) -> QGroupBox:
        grupo = QGroupBox("5. Ruta ICOCIV equivalente")
        form = QFormLayout(grupo)
        self.lbl_ruta_icociv = QLabel("Ruta ICOCIV: cargue un archivo y seleccione la jerarquía.")
        self.lbl_ruta_icociv.setWordWrap(True)
        self.lbl_ruta_icociv.setToolTip(AYUDAS["ruta_icociv"])
        self.combo_grupo_icociv = self._combo_selector(lambda: self._reiniciar_desde_nivel2())
        self.chk_t16 = QCheckBox(texto_checkbox("t16_costos"))
        self.chk_t16.stateChanged.connect(lambda _: self._reiniciar_desde_nivel2())
        self.chk_t16_1 = QCheckBox(texto_checkbox("subclase_costos"))
        self.chk_t16_1.stateChanged.connect(lambda _: self._reiniciar_desde_nivel3())
        self.chk_t16_2 = QCheckBox(texto_checkbox("tipologia_costos"))
        self.chk_t16_2.stateChanged.connect(lambda _: self._reiniciar_desde_nivel4())
        self.chk_t16_3 = QCheckBox(texto_checkbox("capitulo_costos"))
        self.chk_t16_3.stateChanged.connect(lambda _: self._reiniciar_desde_nivel5())
        self.combo_nivel2 = self._combo_selector(lambda: self._reiniciar_desde_nivel3())
        self.combo_nivel3 = self._combo_selector(lambda: self._reiniciar_desde_nivel4())
        self.combo_nivel4 = self._combo_selector(lambda: self._reiniciar_desde_nivel5())
        self.combo_nivel5 = self._combo_selector(lambda: self._reiniciar_desde_nivel6())
        self.combo_nivel6 = self._combo_selector(lambda: self._actualizar_jerarquia_icociv())
        self.lbl_nivel2 = QLabel("Nivel 2")
        self.lbl_nivel3 = QLabel("Nivel 3")
        self.lbl_nivel4 = QLabel("Nivel 4")
        self.lbl_nivel5 = QLabel("Nivel 5")
        self.lbl_nivel6 = QLabel("Nivel 6")
        form.addRow(self.lbl_ruta_icociv)
        _fila(form, nombre_nivel("grupo_obra"), self.combo_grupo_icociv, AYUDAS["ruta_icociv"])
        form.addRow(self.chk_t16)
        _fila(form, self.lbl_nivel2, self.combo_nivel2, "Segundo nivel de la jerarquía ICOCIV disponible para el grupo seleccionado.")
        form.addRow(self.chk_t16_1)
        _fila(form, self.lbl_nivel3, self.combo_nivel3, "Tercer nivel de la jerarquía ICOCIV, según la rama seleccionada.")
        form.addRow(self.chk_t16_2)
        _fila(form, self.lbl_nivel4, self.combo_nivel4, "Cuarto nivel de la jerarquía ICOCIV, si aplica.")
        form.addRow(self.chk_t16_3)
        _fila(form, self.lbl_nivel5, self.combo_nivel5, "Quinto nivel de la jerarquía ICOCIV, si aplica.")
        _fila(form, self.lbl_nivel6, self.combo_nivel6, "Nivel de insumo cuando existe desagregación suficiente.")
        self._ocultar_niveles()
        return grupo

    def _grupo_resultado(self) -> QGroupBox:
        grupo = QGroupBox("8. Resultado del cálculo")
        layout = QVBoxLayout(grupo)
        self.resultado = QTextBrowser()
        self.resultado.setMinimumHeight(220)
        self.resultado.setHtml("<p>Ejecute un cálculo para ver I, I0, R1, R2, explicaciones y trazabilidad.</p>")
        layout.addWidget(self.resultado)
        return grupo

    def _grupo_tabla(self) -> QGroupBox:
        grupo = QGroupBox("9. Tablas del análisis")
        layout = QVBoxLayout(grupo)

        layout.addWidget(QLabel("1. Equivalencias ICCP-ICOCIV"))
        self.tabla_equivalencias = QTableView()
        self.tabla_equivalencias.setModel(self.modelo_equivalencias)
        self.tabla_equivalencias.setAlternatingRowColors(True)
        self.tabla_equivalencias.setMinimumHeight(130)
        self.tabla_equivalencias.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.tabla_equivalencias, 1)

        layout.addWidget(QLabel("2. Cálculo del valor ajustado"))
        self.tabla = QTableView()
        self.tabla.setModel(self.modelo_valor_ajustado)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setMinimumHeight(170)
        self.tabla.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.tabla, 1)

        acciones = QHBoxLayout()
        self.boton_detalle = QPushButton("Ver detalle seleccionado")
        self.boton_editar = QPushButton("Editar seleccionado")
        self.boton_eliminar = QPushButton("Eliminar seleccionado")
        self.boton_exportar = QPushButton("Exportar Excel")
        self.boton_informe = QPushButton("Generar informe")
        self.boton_informe.setToolTip(
            "Informe de ajuste con fórmulas, sustitución numérica y advertencias contractuales."
        )
        self.boton_detalle.clicked.connect(self.ver_detalle_seleccionado)
        self.boton_editar.clicked.connect(self.editar_seleccionado)
        self.boton_eliminar.clicked.connect(self.eliminar_seleccionado)
        self.boton_exportar.clicked.connect(self.exportar_excel)
        self.boton_informe.clicked.connect(self.generar_informe)
        acciones.addWidget(self.boton_detalle)
        acciones.addWidget(self.boton_editar)
        acciones.addWidget(self.boton_eliminar)
        acciones.addWidget(self.boton_exportar)
        acciones.addWidget(self.boton_informe)
        acciones.addStretch()
        layout.addLayout(acciones)
        return grupo

    def ver_detalle_seleccionado(self) -> None:
        indice = self._fila_seleccionada()
        if indice is not None:
            self._mostrar_resultado(self.calculos[indice])

    def limpiar_formulario(self) -> None:
        self.indice_edicion = None
        self.item.clear()
        self.codigo.clear()
        self.unidad.setCurrentIndex(0)
        self.cantidad.setValue(0)
        self.precio_base.setValue(0)
        self.anticipo.setValue(0)
        self.combo_tipo_iccp.setCurrentIndex(0)
        self._actualizar_series_iccp()
        self.observacion_item.clear()
        self.chk_acero.setChecked(False)
        self.p0.setValue(0)
        self.ix.setValue(0)
        self.q.setValue(0)
        self.boton_calcular.setText("Calcular actualización")

    def calcular(self) -> None:
        try:
            try:
                opcion = self._opcion_icociv_actual()
            except ValueError:
                opcion = {"ruta": "", "ruta_estructurada": [], "indices": {}, "seleccion": {}}
            unidad = self.unidad.currentData()
            entrada = {
                "item": self.item.text().strip(),
                "codigo_item": self.codigo.text().strip(),
                "unidad": str(unidad or ""),
                "cantidad": self.cantidad.value() or None,
                "precio_base": self.precio_base.value(),
                "anticipo_amortizado": self.anticipo.value(),
                "fecha_inicial": self._periodo_inicial(),
                "fecha_final": self._periodo_final(),
                **self._serie_iccp_actual(),
                "ruta_icociv": opcion["ruta"],
                "observacion_tecnica": self.observacion_item.toPlainText().strip(),
                "calculo_acero": self.chk_acero.isChecked(),
                "p0": self.p0.value(),
                "ix": self.ix.value(),
                "q": self.q.value(),
            }
            if not entrada["item"]:
                raise ValueError("Debe ingresar el insumo contractual antes de calcular.")
            if not unidad:
                raise ValueError("Seleccione una unidad válida.")
            opcion = self._preparar_icociv_para_empalme(opcion, entrada["fecha_final"])
            resultado = calcular_empalme_iccp_icociv(entrada, opcion["indices"])
        except ValueError as exc:
            QMessageBox.warning(self, "Validación empalme", str(exc))
            return

        resultado.update(self._datos_generales())
        resultado["seleccion_icociv"] = opcion["seleccion"]
        resultado["ruta_icociv_estructurada"] = opcion["ruta_estructurada"]
        resultado.update(opcion.get("metadata_proyeccion") or {})
        if self.indice_edicion is None:
            self.calculos.append(resultado)
        else:
            resultado["calculo_id"] = self.calculos[self.indice_edicion].get("calculo_id")
            self.calculos[self.indice_edicion] = resultado
            self.indice_edicion = None
            self.boton_calcular.setText("Calcular actualización")
        self._renumerar()
        self._mostrar_resultado(resultado)
        self._actualizar_tabla()

    def editar_seleccionado(self) -> None:
        indice = self._fila_seleccionada()
        if indice is None:
            return
        r = self.calculos[indice]
        self.indice_edicion = indice
        self.item.setText(str(r.get("item", "")))
        self.codigo.setText(str(r.get("codigo_item", "")))
        self._establecer_combo_por_data(self.unidad, r.get("unidad"))
        self.cantidad.setValue(float(r.get("cantidad") or 0))
        self.precio_base.setValue(float(r.get("precio_base") or 0))
        self.anticipo.setValue(float(r.get("anticipo_amortizado") or 0))
        self._establecer_periodo(self.fecha_inicial_anio, self.fecha_inicial_mes, r.get("fecha_inicial"))
        self._establecer_periodo(self.fecha_final_anio, self.fecha_final_mes, r.get("fecha_final"))
        self._restaurar_iccp(r)
        self._restaurar_seleccion_icociv(r.get("seleccion_icociv") or {})
        self.observacion_item.setPlainText(str(r.get("observacion_tecnica", "")))
        es_acero = str(r.get("tipo_calculo", "")).lower().endswith("acero")
        self.chk_acero.setChecked(es_acero)
        self.p0.setValue(float(r.get("p0") or 0))
        self.ix.setValue(float(r.get("ix") or 0))
        self.q.setValue(float(r.get("q") or 0))
        self.boton_calcular.setText("Actualizar cálculo")

    def eliminar_seleccionado(self) -> None:
        indice = self._fila_seleccionada()
        if indice is None:
            return
        del self.calculos[indice]
        self.indice_edicion = None
        self.boton_calcular.setText("Calcular actualización")
        self._renumerar()
        self._actualizar_tabla()

    def exportar_excel(self) -> None:
        if not self.calculos:
            QMessageBox.information(self, "Exportación", "No hay cálculos para exportar.")
            return
        errores = _errores_exportacion_excel(self.calculos)
        if errores:
            QMessageBox.warning(self, "Exportación", "No se puede exportar:\n- " + "\n- ".join(errores[:8]))
            return
        ruta, _ = QFileDialog.getSaveFileName(self, "Exportar cálculos de empalme", "calculos_empalme_iccp_icociv.xlsx", "Excel (*.xlsx)")
        if not ruta:
            return
        if not ruta.lower().endswith(".xlsx"):
            ruta = f"{ruta}.xlsx"
        try:
            self._generar_excel_empalme(ruta)
        except Exception as exc:  # pragma: no cover - mensaje de UI
            QMessageBox.critical(self, "Exportación", f"No fue posible generar el archivo:\n{exc}")
            return
        QMessageBox.information(self, "Exportación", f"Archivo generado: {Path(ruta).name}")

    def generar_informe(self) -> None:
        """Informe de ajuste ICCP-ICOCIV en DOCX o PDF, con selector previo."""
        from app_icociv.config.rutas import CARPETA_REPORTES, asegurar_carpeta
        from app_icociv.interfaz.widgets.dialogo_informe import pedir_configuracion
        from app_icociv.reportes.generador_reportes import generar_informe_empalme
        from app_icociv.reportes.modelo import nombre_archivo_informe

        if not self.calculos:
            QMessageBox.information(self, "Informe", "No hay cálculos para incluir en el informe.")
            return
        configuracion = pedir_configuracion(self, tipo_inicial="empalme", formato="DOCX/PDF", solo_empalme=True)
        if configuracion is None:
            return

        generales = self._datos_generales()
        referencia = configuracion.institucional.contrato or generales.get("contrato") or "Ajuste"
        sugerida = asegurar_carpeta(CARPETA_REPORTES) / nombre_archivo_informe("empalme", referencia, "docx")
        ruta, filtro = QFileDialog.getSaveFileName(
            self, "Guardar informe de ajuste", str(sugerida),
            "Documento Word (*.docx);;Documento PDF (*.pdf)",
        )
        if not ruta:
            return
        formato = "pdf" if ruta.lower().endswith(".pdf") or "pdf" in filtro.lower() else "docx"
        try:
            destino = generar_informe_empalme(ruta, self.calculos, generales, configuracion, formato)
        except Exception as exc:  # pragma: no cover - mensaje de UI
            QMessageBox.critical(self, "Informe", f"No fue posible generar el informe:\n{exc}")
            return
        QMessageBox.information(self, "Informe", f"Informe generado: {Path(destino).name}")

    def _generar_excel_empalme(self, ruta: str) -> None:
        """Exporta todos los cálculos y fórmulas en una sola hoja verificable."""
        wb = Workbook()
        ws = wb.active
        ws.title = HOJA_EXPORTACION_EMPALME
        _activar_recalculo_excel(wb)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
        celda = ws.cell(row=1, column=1, value="ACTUALIZACIÓN DE PRECIOS CON EMPALME ICCP-ICOCIV")
        celda.font = Font(bold=True, size=14, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F2937")
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26
        ws.freeze_panes = "A2"

        fila = 3
        fila = _escribir_pares_excel(ws, fila, "Información general", self._filas_info_general_excel())
        fila = _escribir_tabla_excel(ws, fila + 1, "Trazabilidad del cálculo", self._columnas_trazabilidad_excel(), self._filas_trazabilidad_excel())
        fila = _escribir_tabla_excel(
            ws,
            fila + 1,
            "Equivalencias ICCP-ICOCIV",
            [titulo for _, titulo in COLUMNAS_EQUIVALENCIAS],
            [list(self._fila_equivalencia(r).values()) for r in self.calculos],
        )
        fila = self._escribir_calculo_valor_ajustado_excel(ws, fila + 1)
        fila = _escribir_pares_excel(ws, fila + 1, "Metodología general ICCP-ICOCIV", self._filas_metodologia_general_excel())
        if any(r.get("tipo_calculo") == "Cálculo especial acero" for r in self.calculos):
            fila = _escribir_pares_excel(ws, fila, "Metodología especial acero", self._filas_metodologia_acero_excel())
            fila = self._escribir_detalle_acero_excel(ws, fila)
        _escribir_pares_excel(ws, fila, "Observaciones / notas", self._filas_notas_excel())
        _ajustar_hoja_excel(ws)
        Path(ruta).parent.mkdir(parents=True, exist_ok=True)
        wb.save(ruta)

    def _filas_info_general_excel(self) -> list[tuple[str, Any]]:
        datos = self._datos_generales()
        primero = self.calculos[0]
        ultimo = self.calculos[-1]
        return [
            ("Nombre del contrato", datos.get("contrato") or "No registrado"),
            ("Objeto / descripción", datos.get("objeto_contrato") or "No registrado"),
            ("Responsable técnico", datos.get("responsable_tecnico") or "No registrado"),
            ("Fecha inicial del análisis", _periodo_visible(primero.get("fecha_inicial"))),
            ("Fecha final del análisis", _periodo_visible(ultimo.get("fecha_final"))),
            ("Fecha de exportación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Archivo ICOCIV usado", str(self.controlador_icociv.ruta_archivo or "No registrado")),
            ("Fuente ICCP", "Anexo 10 histórico"),
            ("Observación general", datos.get("observacion_general") or "No registrado"),
        ]

    @staticmethod
    def _columnas_trazabilidad_excel() -> list[str]:
        return [
            "ID CÁLCULO",
            "INSUMO / ÍTEM",
            "FECHA INICIAL",
            "FECHA FINAL",
            "TIPO SERIE ICCP",
            "SERIE ICCP USADA",
            "RUTA ICCP COMPLETA",
            "RUTA ICOCIV COMPLETA",
            "ÍNDICE ICCP INICIAL (I0)",
            "ÍNDICE ICCP FINAL (I)",
            "ÍNDICE ICOCIV INICIAL (I0)",
            "ÍNDICE ICOCIV FINAL (I)",
            "ÍNDICE ICOCIV FINAL PROYECTADO",
            "MODELO DE PROYECCIÓN",
            "ESTADO DEL HORIZONTE",
            "OBSERVACIÓN TÉCNICA",
            "FECHA/HORA DEL CÁLCULO",
        ]

    def _filas_trazabilidad_excel(self) -> list[list[Any]]:
        filas = []
        for r in self.calculos:
            proyectado = bool(r.get("icociv_final_es_proyectado"))
            filas.append(
                [
                    r.get("calculo_id") or r.get("numero_calculo") or "",
                    _insumo_contractual(r),
                    _periodo_visible(r.get("fecha_inicial")),
                    _periodo_visible(r.get("fecha_final")),
                    r.get("tipo_serie_iccp_visible") or r.get("tipo_serie_iccp") or "",
                    r.get("serie_iccp") or "",
                    r.get("ruta_iccp") or "",
                    r.get("ruta_icociv") or "",
                    _numero_excel(r.get("i0_iccp")),
                    _numero_excel(r.get("i_iccp")),
                    _numero_excel(r.get("i0_icociv")),
                    _numero_excel(r.get("i_icociv")),
                    "Sí" if proyectado else "No aplica",
                    r.get("modelo_proyeccion") if proyectado else "No aplica",
                    r.get("estado_horizonte") if proyectado else "No aplica",
                    r.get("observacion_tecnica") or "",
                    r.get("fecha_calculo") or "",
                ]
            )
        return filas

    def _escribir_calculo_valor_ajustado_excel(self, ws, fila: int) -> int:
        encabezados = [titulo for _, titulo in self._columnas_valor_ajustado()]
        fila = _escribir_titulo_seccion_excel(ws, fila, "Cálculo del valor ajustado")
        for columna, titulo in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila, column=columna, value=titulo)
            _estilo_encabezado_tabla(celda)
        ws.cell(row=fila, column=12, value="ANTICIPO AMORTIZADO / AUXILIAR")
        _estilo_encabezado_tabla(ws.cell(row=fila, column=12))
        ws.row_dimensions[fila].height = 48
        ws.auto_filter.ref = f"A{fila}:K{fila + len(self.calculos)}"
        self._fila_datos_calculo = fila + 1

        for r in self.calculos:
            fila += 1
            anticipo = 0.0 if r.get("tipo_calculo") == "Cálculo especial acero" else _numero_excel(r.get("anticipo_amortizado")) or 0.0
            base = f"(C{fila}-L{fila})"
            valores = [
                _insumo_contractual(r),
                r.get("unidad") or "",
                _numero_excel(r.get("precio_base")),
                _numero_excel(r.get("i0_iccp")),
                _numero_excel(r.get("i_iccp")),
                f'=IF(OR(D{fila}="",E{fila}=""),0,{base}*((E{fila}/D{fila})-1))',
                _numero_excel(r.get("i0_icociv")),
                _numero_excel(r.get("i_icociv")),
                f'=IF(OR(G{fila}="",H{fila}=""),0,({base}+F{fila})*((H{fila}/G{fila})-1))',
                f"=F{fila}+I{fila}",
                f"={base}+J{fila}",
                anticipo,
            ]
            for columna, valor in enumerate(valores, start=1):
                celda = ws.cell(row=fila, column=columna, value=valor)
                _estilo_celda_tabla(celda)
            for columna in (3, 6, 9, 10, 11, 12):
                ws.cell(row=fila, column=columna).number_format = FORMATO_MONEDA_EXCEL
            for columna in (4, 5, 7, 8):
                ws.cell(row=fila, column=columna).number_format = FORMATO_INDICE_EXCEL
            for columna in (6, 9, 10):
                ws.cell(row=fila, column=columna).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["L"].hidden = True
        return fila + 1

    def _filas_notas_excel(self) -> list[tuple[str, Any]]:
        return [
            ("Metodología", "R = R1 + R2 según los lineamientos de empalme ICCP a ICOCIV."),
            ("Fórmulas verificables", "Las columnas R1, R2, VALOR DEL AJUSTE (R) y VALOR INSUMO AJUSTADO se exportan como fórmulas de Excel."),
            ("Proyecciones", "Si se usó índice ICOCIV proyectado, queda marcado en la trazabilidad del cálculo."),
        ]

    @staticmethod
    def _filas_metodologia_general_excel() -> list[tuple[str, Any]]:
        return [
            ("R1", "R1 = (P - A) × [(I / I0) - 1]"),
            ("R2", "R2 = [(P - A) + R1] × [(I / I0) - 1]"),
            ("R", "R = R1 + R2"),
            ("Valor actualizado", "Valor actualizado = (P - A) + R"),
            ("Variables", "P: valor base. A: anticipo amortizado. I/I0: índices del tramo ICCP o ICOCIV."),
        ]

    @staticmethod
    def _filas_metodologia_acero_excel() -> list[tuple[str, Any]]:
        return [
            ("R1 acero", "R1 = P0 × [(I / I0) - 1]"),
            ("R2 acero", "R2 = (P0 + R1) × [(I / I0) - 1]"),
            ("R acero", "R = R1 + R2"),
            ("Z", "Z = (Ix × q) - (R + P0)"),
            ("Variables", "P0: valor base acero. Ix: valor facturado por kg. q: kg ejecutados. Z: valor adicional por fluctuación."),
        ]

    def _escribir_detalle_acero_excel(self, ws, fila: int) -> int:
        encabezados = ["INSUMO CONTRACTUAL", "P0", "Ix", "q", "Ix × q", "R1", "R2", "R", "P0 + R", "Z", "OBSERVACIÓN"]
        fila = _escribir_titulo_seccion_excel(ws, fila, "Detalle acero")
        for columna, titulo in enumerate(encabezados, start=1):
            celda = ws.cell(row=fila, column=columna, value=titulo)
            _estilo_encabezado_tabla(celda)
        ws.row_dimensions[fila].height = 36
        for indice, r in enumerate(self.calculos, start=1):
            if r.get("tipo_calculo") != "Cálculo especial acero":
                continue
            fila += 1
            # Referencia a la fila del mismo cálculo en la tabla de valor ajustado (misma hoja).
            fila_calculo = self._fila_datos_calculo + indice - 1
            valores = [
                _insumo_contractual(r),
                _numero_excel(r.get("p0")),
                _numero_excel(r.get("ix")),
                _numero_excel(r.get("q")),
                f'=IF(OR(C{fila}="",D{fila}=""),"",C{fila}*D{fila})',
                f"=F{fila_calculo}",
                f"=I{fila_calculo}",
                f"=J{fila_calculo}",
                f"=B{fila}+H{fila}",
                f'=IF(E{fila}="","No calculado",E{fila}-(H{fila}+B{fila}))',
                r.get("z_observacion") or r.get("observacion_tecnica") or "",
            ]
            for columna, valor in enumerate(valores, start=1):
                celda = ws.cell(row=fila, column=columna, value=valor)
                _estilo_celda_tabla(celda)
            for columna in (2, 3, 5, 6, 7, 8, 9, 10):
                ws.cell(row=fila, column=columna).number_format = FORMATO_MONEDA_EXCEL
            ws.cell(row=fila, column=4).number_format = FORMATO_INDICE_EXCEL
        return fila + 1

    def _opcion_icociv_actual(self) -> dict[str, Any]:
        seleccion = self._seleccion_icociv_actual()
        if not self.controlador_icociv.archivo_cargado or seleccion.get("idx_g") is None:
            raise ValueError("Debe cargar un archivo ICOCIV y seleccionar al menos el grupo de obra.")
        fuente, fila = resolver_fila_seleccionada(self.controlador_icociv.tablas, self.controlador_icociv.periodos, seleccion)
        self.controlador_icociv.fuente_actual = fuente
        indices = {
            periodo: float(fila.iloc[0][periodo])
            for periodo in self.controlador_icociv.periodos
            if periodo in fila.columns and pd.notna(fila.iloc[0][periodo])
        }
        if not indices:
            raise ValueError("La ruta ICOCIV seleccionada no tiene índices mensuales disponibles.")
        ruta = self._texto_ruta_icociv(seleccion)
        return {
            "ruta": ruta,
            "ruta_estructurada": self.controlador_icociv.construir_ruta_jerarquica(seleccion),
            "indices": indices,
            "seleccion": seleccion,
        }

    def _preparar_icociv_para_empalme(self, opcion: dict[str, Any], fecha_final: Any) -> dict[str, Any]:
        indices = dict(opcion.get("indices") or {})
        if not indices:
            return opcion

        periodo_final = normalizar_periodo_empalme(fecha_final)
        ultimo_real = _ultimo_periodo(indices)
        if _periodo_orden(periodo_final) <= _periodo_orden(ultimo_real):
            return {**opcion, "indices": indices, "metadata_proyeccion": {"icociv_final_es_proyectado": False}}

        if self.callback_proyeccion_icociv is None:
            raise ValueError(
                "La fecha final solicitada supera el último periodo disponible en el archivo ICOCIV, "
                "pero el módulo de proyección no está conectado."
            )

        QMessageBox.information(
            self,
            "Proyección ICOCIV requerida",
            "La fecha final solicitada supera el último periodo disponible en el archivo ICOCIV. "
            "Se usará el módulo de proyección existente para estimar el índice ICOCIV final.",
        )
        anio, mes = _periodo_partes(periodo_final)
        resultado_ui = self.callback_proyeccion_icociv(opcion["seleccion"], anio, mes)
        proyeccion = resultado_ui.get("proyeccion", {})
        solicitado = proyeccion.get("resultado_horizonte_solicitado") or {}
        indice_final = solicitado.get("indice_proyectado")
        if not solicitado.get("proyeccion_generada") or indice_final is None:
            razon = solicitado.get("razon_principal") or "Revise el estado del horizonte y el máximo recomendado en la pestaña de proyección."
            raise ValueError(
                "No fue posible usar la fecha final solicitada porque el módulo de proyección no considera viable ese horizonte. "
                f"{razon}"
            )

        for fila in _filas_proyeccion(proyeccion):
            periodo = normalizar_periodo_empalme(fila.get("periodo"))
            indices[periodo] = float(fila["indice_proyectado"])
        indices[periodo_final] = float(indice_final)
        metadata = {
            "icociv_final_es_proyectado": True,
            "ultimo_periodo_icociv_real": ultimo_real,
            "indice_icociv_final_proyectado": float(indice_final),
            "modelo_proyeccion": solicitado.get("modelo_aplicado") or proyeccion.get("model_name", ""),
            "horizonte_usado": solicitado.get("horizonte_solicitado") or proyeccion.get("horizonte_solicitado"),
            "estado_horizonte": solicitado.get("estado", ""),
            "periodo_proyectado": solicitado.get("periodo_proyectado") or periodo_final,
            "advertencias_proyeccion": solicitado.get("razones_tecnicas") or [],
        }
        QMessageBox.information(
            self,
            "Índice ICOCIV proyectado",
            "Índice ICOCIV final proyectado usado en el empalme: "
            f"{_fmt_indice(indice_final)}\n"
            f"Periodo proyectado: {_periodo_visible(periodo_final)}\n"
            f"Modelo seleccionado: {metadata['modelo_proyeccion']}\n"
            f"Estado del horizonte: {metadata['estado_horizonte']}",
        )
        return {**opcion, "indices": indices, "metadata_proyeccion": metadata}

    def _actualizar_jerarquia_icociv(self) -> None:
        if not self.controlador_icociv.archivo_cargado:
            return
        seleccion = self._seleccion_icociv_actual()
        estado = self.controlador_icociv.obtener_estado_jerarquia(seleccion)
        self._rellenar_nivel(self.lbl_nivel2, self.combo_nivel2, estado["nivel2"])
        self._rellenar_nivel(self.lbl_nivel3, self.combo_nivel3, estado["nivel3"])
        self._rellenar_nivel(self.lbl_nivel4, self.combo_nivel4, estado["nivel4"])
        self._rellenar_nivel(self.lbl_nivel5, self.combo_nivel5, estado["nivel5"])
        self._rellenar_nivel(self.lbl_nivel6, self.combo_nivel6, estado["nivel6"])
        self._establecer_visible_checkbox(self.chk_t16_1, estado["mostrar_chk_t16_1"])
        self._establecer_visible_checkbox(self.chk_t16_2, estado["mostrar_chk_t16_2"])
        self._establecer_visible_checkbox(self.chk_t16_3, estado["mostrar_chk_t16_3"])
        self.lbl_ruta_icociv.setText(self._texto_ruta_icociv(self._seleccion_icociv_actual()))

    def _texto_ruta_icociv(self, seleccion: dict[str, Any]) -> str:
        try:
            fuente, _ = resolver_fila_seleccionada(self.controlador_icociv.tablas, self.controlador_icociv.periodos, seleccion)
            self.controlador_icociv.fuente_actual = fuente
        except Exception:
            return "Ruta ICOCIV: sin selección válida."
        ruta = self.controlador_icociv.construir_ruta_jerarquica(seleccion)
        ruta_sin_tabla = [item for item in ruta if item.get("nivel") != "Tabla ICOCIV"]
        if not ruta_sin_tabla:
            return "Ruta ICOCIV: sin selección válida."
        return "Ruta ICOCIV: " + " > ".join(item.get("valor", "") for item in ruta_sin_tabla)

    def _seleccion_icociv_actual(self) -> dict[str, Any]:
        return {
            "idx_g": self.combo_grupo_icociv.currentData(),
            "chk_T16": self.chk_t16.isChecked(),
            "idx_l2": self.combo_nivel2.currentData(),
            "idx_l3": self.combo_nivel3.currentData(),
            "idx_l4": self.combo_nivel4.currentData(),
            "idx_l5": self.combo_nivel5.currentData(),
            "idx_l6": self.combo_nivel6.currentData(),
            "chk_T16_1": self.chk_t16_1.isChecked(),
            "chk_T16_2": self.chk_t16_2.isChecked(),
            "chk_T16_3": self.chk_t16_3.isChecked(),
        }

    def _restaurar_seleccion_icociv(self, seleccion: dict[str, Any]) -> None:
        self._establecer_combo_por_data(self.combo_grupo_icociv, seleccion.get("idx_g"))
        self.chk_t16.setChecked(bool(seleccion.get("chk_T16", False)))
        self._actualizar_jerarquia_icociv()
        self._establecer_combo_por_data(self.combo_nivel2, seleccion.get("idx_l2"))
        self.chk_t16_1.setChecked(bool(seleccion.get("chk_T16_1", False)))
        self._actualizar_jerarquia_icociv()
        self._establecer_combo_por_data(self.combo_nivel3, seleccion.get("idx_l3"))
        self.chk_t16_2.setChecked(bool(seleccion.get("chk_T16_2", False)))
        self._actualizar_jerarquia_icociv()
        self._establecer_combo_por_data(self.combo_nivel4, seleccion.get("idx_l4"))
        self.chk_t16_3.setChecked(bool(seleccion.get("chk_T16_3", False)))
        self._actualizar_jerarquia_icociv()
        self._establecer_combo_por_data(self.combo_nivel5, seleccion.get("idx_l5"))
        self._establecer_combo_por_data(self.combo_nivel6, seleccion.get("idx_l6"))
        self._actualizar_jerarquia_icociv()

    def _combo_selector(self, accion_cambio) -> ComboBoxSinRueda:
        combo = ComboBoxSinRueda()
        combo.setMaxVisibleItems(18)
        combo.setMinimumContentsLength(18)
        combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        combo.currentIndexChanged.connect(lambda _: accion_cambio())
        return combo

    def _actualizar_series_iccp(self) -> None:
        tipo = self.combo_tipo_iccp.currentData()
        opciones = [{"texto": nombre, "indice": nombre} for nombre in self.series_iccp.get(tipo, [])]
        self._llenar_combo(self.combo_serie_iccp, opciones)
        self.combo_serie_iccp.setEnabled(bool(tipo))

    def _serie_iccp_actual(self) -> dict[str, str]:
        tipo = self.combo_tipo_iccp.currentData()
        serie = self.combo_serie_iccp.currentData()
        if not tipo:
            raise ValueError("Seleccione el tipo de serie ICCP.")
        if not serie:
            raise ValueError("Seleccione la serie ICCP.")
        if serie not in self.series_iccp.get(tipo, []):
            raise ValueError("La serie ICCP seleccionada no corresponde al tipo elegido.")
        return {
            "tipo_serie_iccp": str(tipo or ""),
            "serie_iccp": str(serie or ""),
        }

    def _periodo_inicial(self) -> str:
        return f"{self.fecha_inicial_anio.value()}_{self.fecha_inicial_mes.value()}"

    def _periodo_final(self) -> str:
        return f"{self.fecha_final_anio.value()}_{self.fecha_final_mes.value()}"

    @staticmethod
    def _establecer_periodo(spin_anio: QSpinBox, spin_mes: QSpinBox, periodo: Any) -> None:
        partes = str(periodo or "").replace("-", "_").split("_")
        if len(partes) >= 2 and partes[0].isdigit() and partes[1].isdigit():
            spin_anio.setValue(int(partes[0]))
            spin_mes.setValue(int(partes[1]))

    def _restaurar_iccp(self, resultado: dict[str, Any]) -> None:
        self._establecer_combo_por_data(self.combo_tipo_iccp, resultado.get("tipo_serie_iccp"))
        self._actualizar_series_iccp()
        self._establecer_combo_por_data(self.combo_serie_iccp, resultado.get("serie_iccp"))

    def _rellenar_nivel(self, etiqueta: QLabel, combo: QComboBox, descriptor: dict[str, Any] | None) -> None:
        valor_actual = combo.currentData()
        if not descriptor or not descriptor["opciones"]:
            self._llenar_combo(combo, [], None)
            etiqueta.setVisible(False)
            combo.parentWidget().setVisible(False)
            return
        etiqueta.setText(descriptor["etiqueta"])
        etiqueta.setVisible(True)
        combo.parentWidget().setVisible(True)
        self._llenar_combo(combo, descriptor["opciones"], valor_actual)

    def _llenar_combo(self, combo: QComboBox, opciones: list[dict[str, Any]], valor_actual: Any = None) -> None:
        combo.blockSignals(True)
        combo.clear()
        combo.addItem("Sin selección", None)
        indice_restaurar = 0
        for opcion in opciones:
            combo.addItem(opcion["texto"], opcion["indice"])
            combo.setItemData(combo.count() - 1, opcion["texto"], Qt.ItemDataRole.ToolTipRole)
            if opcion["indice"] == valor_actual:
                indice_restaurar = combo.count() - 1
        combo.setCurrentIndex(indice_restaurar)
        combo.blockSignals(False)

    def _reiniciar_desde_nivel2(self) -> None:
        self._limpiar_combos(self.combo_nivel2, self.combo_nivel3, self.combo_nivel4, self.combo_nivel5, self.combo_nivel6)
        self._actualizar_jerarquia_icociv()

    def _reiniciar_desde_nivel3(self) -> None:
        self._limpiar_combos(self.combo_nivel3, self.combo_nivel4, self.combo_nivel5, self.combo_nivel6)
        self._actualizar_jerarquia_icociv()

    def _reiniciar_desde_nivel4(self) -> None:
        self._limpiar_combos(self.combo_nivel4, self.combo_nivel5, self.combo_nivel6)
        self._actualizar_jerarquia_icociv()

    def _reiniciar_desde_nivel5(self) -> None:
        self._limpiar_combos(self.combo_nivel5, self.combo_nivel6)
        self._actualizar_jerarquia_icociv()

    def _reiniciar_desde_nivel6(self) -> None:
        self._limpiar_combos(self.combo_nivel6)
        self._actualizar_jerarquia_icociv()

    def _limpiar_combos(self, *combos: QComboBox) -> None:
        for combo in combos:
            self._llenar_combo(combo, [])

    def _ocultar_niveles(self) -> None:
        for widget in (self.lbl_nivel2, self.lbl_nivel3, self.lbl_nivel4, self.lbl_nivel5, self.lbl_nivel6):
            widget.setVisible(False)
        for combo in (self.combo_nivel2, self.combo_nivel3, self.combo_nivel4, self.combo_nivel5, self.combo_nivel6):
            combo.parentWidget().setVisible(False)
        for checkbox in (self.chk_t16_1, self.chk_t16_2, self.chk_t16_3):
            checkbox.setVisible(False)

    def _establecer_visible_checkbox(self, checkbox: QCheckBox, visible: bool) -> None:
        if not visible:
            checkbox.blockSignals(True)
            checkbox.setChecked(False)
            checkbox.blockSignals(False)
        checkbox.setVisible(visible)

    def _actualizar_controles_selector(self, habilitado: bool) -> None:
        for widget in (self.combo_grupo_icociv, self.chk_t16, self.combo_nivel2, self.combo_nivel3, self.combo_nivel4, self.combo_nivel5, self.combo_nivel6):
            widget.setEnabled(habilitado)

    def _datos_generales(self) -> dict[str, str]:
        return {
            "contrato": self.contrato.text().strip(),
            "objeto_contrato": self.objeto.toPlainText().strip(),
            "responsable_tecnico": self.responsable.text().strip(),
            "observacion_general": self.observacion_general.toPlainText().strip(),
        }

    def _mostrar_resultado(self, r: dict[str, Any]) -> None:
        filas = [
            ("Precio base P", _fmt_moneda(r.get("precio_base")), AYUDAS["precio"]),
            ("Unidad", r.get("unidad"), AYUDAS["unidad"]),
            ("Anticipo amortizado A", _fmt_moneda(r.get("anticipo_amortizado")), AYUDAS["anticipo"]),
            ("Base ajustable P - A", _fmt_moneda(r.get("base_ajustable")), AYUDAS["base"]),
            ("Fecha inicial", r.get("fecha_inicial"), AYUDAS["fecha_inicial"]),
            ("Fecha final", r.get("fecha_final"), AYUDAS["fecha_final"]),
            ("Tipo de serie ICCP", r.get("tipo_serie_iccp_visible"), AYUDAS["tipo_serie_iccp"]),
            ("Serie ICCP usada", r.get("serie_iccp"), AYUDAS["serie_iccp"]),
            ("Ruta ICCP", r.get("ruta_iccp"), AYUDAS["grupo_iccp"]),
            ("I0 ICCP", _fmt_indice(r.get("i0_iccp")), "Índice ICCP inicial correspondiente a la fecha inicial seleccionada."),
            ("I ICCP", _fmt_indice(r.get("i_iccp")), "Índice ICCP final del primer tramo; normalmente diciembre de 2021 cuando cruza la transición."),
            ("Factor ICCP", _fmt_factor(r.get("factor_iccp")), "Relación I ICCP / I0 ICCP usada para calcular R1."),
            ("R1", _fmt_moneda(r.get("r1")), "Primer valor parcial ajustado según lineamientos ICCP-ICOCIV."),
            ("Ruta ICOCIV", r.get("ruta_icociv"), AYUDAS["ruta_icociv"]),
            ("I0 ICOCIV", _fmt_indice(r.get("i0_icociv")), "Índice ICOCIV inicial del segundo tramo; normalmente diciembre de 2021."),
            ("I ICOCIV", _fmt_indice_proyectado(r), "Índice ICOCIV final correspondiente a la fecha final seleccionada."),
            ("Factor ICOCIV", _fmt_factor(r.get("factor_icociv")), "Relación I ICOCIV / I0 ICOCIV usada para calcular R2."),
            ("R2", _fmt_moneda(r.get("r2")), "Segundo valor parcial ajustado según lineamientos ICCP-ICOCIV."),
            ("R total", _fmt_moneda(r.get("r_total")), "Valor definitivo del ajuste: R = R1 + R2."),
            ("Valor actualizado", _fmt_moneda(r.get("valor_actualizado")), "Resultado final después de aplicar el empalme ICCP-ICOCIV."),
            ("Diferencia absoluta", _fmt_moneda(r.get("diferencia_absoluta")), "Diferencia en pesos entre el valor actualizado y la base ajustable inicial."),
            ("Diferencia porcentual", _fmt_porcentaje(r.get("diferencia_porcentual")), "Porcentaje de variación frente a la base ajustable inicial."),
        ]
        if r.get("tipo_calculo") == "Cálculo especial acero":
            filas.extend(
                [
                    ("P0", _fmt_moneda(r.get("p0")), AYUDAS["p0"]),
                    ("Ix", _fmt_moneda(r.get("ix")), AYUDAS["ix"]),
                    ("q", _fmt_indice(r.get("q")), AYUDAS["q"]),
                    ("Valor facturado total Ix × q", _fmt_moneda_o_no_calculado(r.get("valor_facturado_total")), "Valor facturado total del acero. Se calcula como Ix × q."),
                    ("Z", _fmt_moneda_o_no_calculado(r.get("z")), "Valor adicional por fluctuación del acero: compara el valor facturado contra P0 + R."),
                ]
            )
        cuerpo = "".join(
            f"<tr><th>{escape(str(k))}</th><td>{escape('' if v is None else str(v))}</td><td>{escape(str(a))}</td></tr>"
            for k, v, a in filas
        )
        advertencias = "".join(f"<li>{escape(a)}</li>" for a in self._advertencias(r))
        acero = ""
        if r.get("tipo_calculo") == "Cálculo especial acero":
            acero = (
                "<h4>Resultado especial acero</h4>"
                "<table border='1' cellspacing='0' cellpadding='4'>"
                f"<tr><th>Valor facturado total Ix × q</th><td>{escape(_fmt_moneda_o_no_calculado(r.get('valor_facturado_total')))}</td></tr>"
                f"<tr><th>Ajuste R</th><td>{escape(_fmt_moneda(r.get('r_total')))}</td></tr>"
                f"<tr><th>Z</th><td>{escape(_fmt_moneda_o_no_calculado(r.get('z')))}</td></tr>"
                "</table>"
                "<h4>Fórmula especial de acero</h4>"
                "<pre>R1 = P0 × [(I ICCP / I0 ICCP) - 1]\n"
                "R2 = (P0 + R1) × [(I ICOCIV / I0 ICOCIV) - 1]\n"
                "R = R1 + R2\n"
                "Z = (Ix × q) - (R + P0)</pre>"
                "<p><b>P0</b>: valor base del acero. <b>Ix</b>: valor facturado por kg. "
                "<b>q</b>: kg ejecutados. <b>Z</b>: valor adicional por fluctuación.</p>"
            )
        self.resultado.setHtml(
            "<h3>Resultado de actualización</h3>"
            "<table border='1' cellspacing='0' cellpadding='8'>"
            f"<tr><th>Valor actualizado</th><td><b>{escape(_fmt_moneda(r.get('valor_actualizado')))}</b></td></tr>"
            f"<tr><th>Valor del ajuste (R)</th><td>{escape(_fmt_moneda(r.get('r_total')))}</td></tr>"
            f"<tr><th>Índice ICOCIV final usado</th><td>{escape(_fmt_indice_proyectado(r))}</td></tr>"
            f"<tr><th>Diferencia porcentual</th><td>{escape(_fmt_porcentaje(r.get('diferencia_porcentual')))}</td></tr>"
            f"<tr><th>Base ajustable</th><td>{escape(_fmt_moneda(r.get('base_ajustable')))}</td></tr>"
            "</table>"
            f"{_nota_proyeccion_html(r)}"
            "<h4>Detalle del cálculo</h4>"
            "<table border='1' cellspacing='0' cellpadding='4'>"
            "<tr><th>Campo</th><th>Valor</th><th>Qué significa</th></tr>"
            f"{cuerpo}</table>"
            f"{acero}"
            f"<h4>Advertencias</h4><ul>{advertencias}</ul>"
            "<h4>Desarrollo matemático</h4>"
            f"<pre>{escape(_trazabilidad_lineamientos(r))}</pre>"
        )

    def _advertencias(self, r: dict[str, Any]) -> list[str]:
        advertencias = ["La equivalencia ICCP-ICOCIV es una selección manual del usuario y requiere justificación técnica."]
        if r.get("caso") == "solo_iccp":
            advertencias.append("El cálculo usa solo ICCP porque ambas fechas son hasta diciembre de 2021.")
        if r.get("caso") == "solo_icociv":
            advertencias.append("El cálculo usa solo ICOCIV porque ambas fechas son posteriores a diciembre de 2021.")
        if r.get("ruta_icociv") and str(r.get("ruta_icociv")).count(" > ") < 2:
            advertencias.append("La ruta ICOCIV seleccionada parece general; revise si el nivel es suficiente.")
        if r.get("tipo_calculo") == "Cálculo especial acero":
            if r.get("z") is None:
                advertencias.append(r.get("z_observacion") or "Para calcular Z debe ingresar Ix y q.")
            elif r.get("z") <= 0:
                advertencias.append("El valor Z del acero es negativo o cero.")
        return advertencias

    def _actualizar_tabla(self) -> None:
        self.modelo_equivalencias.establecer_dataframe(self._dataframe_equivalencias())
        self.modelo_valor_ajustado.establecer_dataframe(self._dataframe_valor_ajustado())

    def _dataframe_equivalencias(self) -> pd.DataFrame:
        return _dataframe_desde_registros(
            [self._fila_equivalencia(r) for r in self.calculos],
            COLUMNAS_EQUIVALENCIAS,
        )

    def _dataframe_valor_ajustado(self) -> pd.DataFrame:
        return _dataframe_desde_registros(
            [self._fila_valor_ajustado(r) for r in self.calculos],
            self._columnas_valor_ajustado(),
        )

    def _dataframe_detalle_exportacion(self) -> pd.DataFrame:
        registros = []
        for r in self.calculos:
            fila = self._fila_valor_ajustado(r)
            fila.update({clave: _valor_detalle_exportacion(clave, r.get(clave)) for clave, _ in COLUMNAS_DETALLE_EXPORTACION})
            registros.append(fila)
        return _dataframe_desde_registros(registros, self._columnas_valor_ajustado() + COLUMNAS_DETALLE_EXPORTACION)

    def _columnas_valor_ajustado(self) -> list[tuple[str, str]]:
        if not self.calculos:
            return COLUMNAS_VALOR_AJUSTADO
        r = self.calculos[-1]
        fechas = _fechas_columnas_indices(r)
        return [
            (clave, _titulo_indice(clave, titulo, fechas.get(clave), bool(r.get("icociv_final_es_proyectado"))))
            for clave, titulo in COLUMNAS_VALOR_AJUSTADO
        ]

    def _fila_equivalencia(self, r: dict[str, Any]) -> dict[str, Any]:
        return {
            "insumo_contractual": _insumo_contractual(r),
            "grupo_iccp_equivalente": _equivalente_iccp(r),
            "insumo_icociv_equivalente": _equivalente_icociv(r),
        }

    def _fila_valor_ajustado(self, r: dict[str, Any]) -> dict[str, Any]:
        return {
            "insumo_contractual": _insumo_contractual(r),
            "unidad": r.get("unidad", ""),
            "precio_base": _fmt_moneda(r.get("precio_base")),
            "i0_iccp": _fmt_indice_o_na(r.get("i0_iccp")),
            "i_iccp": _fmt_indice_o_na(r.get("i_iccp")),
            "r1": _fmt_moneda_o_na(r.get("r1")),
            "i0_icociv": _fmt_indice_o_na(r.get("i0_icociv")),
            "i_icociv": _fmt_indice_proyectado(r),
            "r2": _fmt_moneda_o_na(r.get("r2")),
            "r_total": _fmt_moneda_o_na(r.get("r_total")),
            "valor_actualizado": _fmt_moneda(r.get("valor_actualizado")),
        }

    def _fila_seleccionada(self) -> int | None:
        seleccion = self.tabla.selectionModel().selectedRows()
        if not seleccion:
            QMessageBox.information(self, "Tabla de cálculos", "Seleccione un cálculo en la tabla.")
            return None
        return int(seleccion[0].row())

    def _alternar_acero(self, activo: bool) -> None:
        for widget in (self.p0, self.ix, self.q):
            widget.setEnabled(activo)

    def _renumerar(self) -> None:
        for numero, calculo in enumerate(self.calculos, start=1):
            calculo["numero_calculo"] = numero
            if not calculo.get("calculo_id"):
                calculo["calculo_id"] = f"calc-{uuid4().hex[:8]}"

    @staticmethod
    def _seleccionar_texto(combo: QComboBox, texto: str) -> None:
        for indice in range(combo.count()):
            if combo.itemText(indice) == texto:
                combo.setCurrentIndex(indice)
                return

    @staticmethod
    def _establecer_combo_por_data(combo: QComboBox, valor: Any) -> None:
        for indice in range(combo.count()):
            if combo.itemData(indice) == valor:
                combo.setCurrentIndex(indice)
                return


def _fila(form: QFormLayout, etiqueta: str | QLabel, widget: QWidget, ayuda: str) -> None:
    form.addRow(_etiqueta_ayuda(etiqueta, ayuda), _ayuda(widget, ayuda))


def _etiqueta_ayuda(etiqueta: str | QLabel, ayuda: str) -> QWidget:
    if isinstance(etiqueta, QLabel):
        etiqueta.setToolTip(ayuda)
        return etiqueta
    contenedor = QWidget()
    layout = QHBoxLayout(contenedor)
    layout.setContentsMargins(0, 0, 0, 0)
    texto = QLabel(str(etiqueta))
    texto.setToolTip(ayuda)
    layout.addWidget(texto)
    layout.addStretch()
    return contenedor


def _ayuda(widget: QWidget, texto: str) -> QWidget:
    contenedor = QWidget()
    layout = QHBoxLayout(contenedor)
    layout.setContentsMargins(0, 0, 0, 0)
    widget.setToolTip(texto)
    layout.addWidget(widget)
    return contenedor


def _periodo_spins(anio: int, mes: int) -> tuple[QSpinBox, QSpinBox]:
    spin_anio = SpinEnteroSinRueda()
    spin_anio.setRange(2000, 2100)
    spin_anio.setValue(anio)
    spin_mes = SpinEnteroSinRueda()
    spin_mes.setRange(1, 12)
    spin_mes.setValue(mes)
    for spin in (spin_anio, spin_mes):
        spin.setAccelerated(True)
        spin.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.NoButtons)
        spin.setMinimumHeight(ALTURA_CONTROL)
        spin.setToolTip("Seleccione año y mes; no permite texto libre.")
    return spin_anio, spin_mes


def _periodo_widget(spin_anio: QSpinBox, spin_mes: QSpinBox) -> QWidget:
    contenedor = QWidget()
    layout = QHBoxLayout(contenedor)
    layout.setContentsMargins(0, 0, 0, 0)
    layout.addWidget(QLabel("Año"))
    layout.addWidget(_control_incremental(spin_anio))
    layout.addWidget(QLabel("Mes"))
    layout.addWidget(_control_incremental(spin_mes))
    layout.addStretch()
    return contenedor


def _control_incremental(spin: QSpinBox) -> QWidget:
    contenedor = QWidget()
    layout = QHBoxLayout(contenedor)
    layout.setContentsMargins(0, 0, 0, 0)
    layout.setSpacing(4)
    boton_menos = _boton_incremento("-")
    boton_mas = _boton_incremento("+")
    boton_menos.clicked.connect(spin.stepDown)
    boton_mas.clicked.connect(spin.stepUp)
    layout.addWidget(boton_menos)
    layout.addWidget(spin, 1)
    layout.addWidget(boton_mas)
    return contenedor


def _boton_incremento(texto: str) -> QPushButton:
    boton = QPushButton(texto)
    boton.setObjectName("boton_incremento")
    boton.setFixedWidth(ANCHO_BOTON_INCREMENTO)
    boton.setMinimumHeight(ALTURA_CONTROL)
    boton.setFocusPolicy(Qt.FocusPolicy.NoFocus)
    return boton


def _spin(minimo: float, maximo: float, decimales: int) -> SpinSinRueda:
    spin = SpinSinRueda()
    spin.setRange(minimo, maximo)
    spin.setDecimals(decimales)
    spin.setSingleStep(1)
    spin.setGroupSeparatorShown(True)
    return spin


def _fmt_col(valor: Any, decimales: int) -> str:
    if valor is None or valor == "":
        return ""
    try:
        texto = f"{float(valor):,.{decimales}f}"
    except (TypeError, ValueError):
        return str(valor)
    return texto.replace(",", "_").replace(".", ",").replace("_", ".")


def _fmt_moneda(valor: Any) -> str:
    texto = _fmt_col(valor, 2)
    return f"$ {texto}" if texto else ""


def _fmt_moneda_o_no_calculado(valor: Any) -> str:
    return _fmt_moneda(valor) or "No calculado. Faltan Ix y/o q."


def _fmt_porcentaje(valor: Any) -> str:
    texto = _fmt_col(valor, 2)
    return f"{texto} %" if texto else ""


def _fmt_indice(valor: Any) -> str:
    return _fmt_col(valor, 4)


def _fmt_factor(valor: Any) -> str:
    return _fmt_col(valor, 6)


def _fmt_indice_o_na(valor: Any) -> str:
    return _fmt_indice(valor) or "No aplica"


def _fmt_moneda_o_na(valor: Any) -> str:
    return _fmt_moneda(valor) or "No aplica"


def _fmt_indice_proyectado(r: dict[str, Any]) -> str:
    texto = _fmt_indice_o_na(r.get("i_icociv"))
    return f"{texto} (proyectado)" if r.get("icociv_final_es_proyectado") and texto != "No aplica" else texto


def _titulo_indice(clave: str, titulo: str, periodo: str | None, proyectado: bool = False) -> str:
    if clave not in {"i0_iccp", "i_iccp", "i0_icociv", "i_icociv"}:
        return titulo
    sufijo = "\n(PROYECTADO)" if clave == "i_icociv" and proyectado else ""
    return f"{titulo}\n{_periodo_visible(periodo)}{sufijo}"


def _fechas_columnas_indices(r: dict[str, Any]) -> dict[str, str | None]:
    caso = r.get("caso")
    return {
        "i0_iccp": r.get("fecha_inicial") if r.get("i0_iccp") is not None else None,
        "i_iccp": (r.get("fecha_final") if caso == "solo_iccp" else "2021_12") if r.get("i_iccp") is not None else None,
        "i0_icociv": (r.get("fecha_inicial") if caso == "solo_icociv" else "2021_12") if r.get("i0_icociv") is not None else None,
        "i_icociv": r.get("fecha_final") if r.get("i_icociv") is not None else None,
    }


def _periodo_visible(periodo: Any) -> str:
    if not periodo:
        return "NO APLICA"
    partes = str(periodo).replace("-", "_").split("_")
    if len(partes) < 2 or not partes[0].isdigit() or not partes[1].isdigit():
        return str(periodo).upper()
    meses = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
    mes = int(partes[1])
    return f"{meses[mes]} DE {partes[0]}" if 1 <= mes <= 12 else str(periodo).upper()


def _periodo_orden(periodo: Any) -> int:
    anio, mes = _periodo_partes(periodo)
    return anio * 12 + mes


def _periodo_partes(periodo: Any) -> tuple[int, int]:
    normalizado = normalizar_periodo_empalme(periodo)
    anio, mes = normalizado.split("_", 1)
    return int(anio), int(mes)


def _ultimo_periodo(indices: dict[str, Any]) -> str:
    periodos = [normalizar_periodo_empalme(periodo) for periodo in indices]
    return max(periodos, key=_periodo_orden)


def _filas_proyeccion(proyeccion: dict[str, Any]) -> list[dict[str, Any]]:
    tabla = proyeccion.get("proyecciones")
    if isinstance(tabla, pd.DataFrame):
        return tabla[["periodo", "indice_proyectado"]].dropna().to_dict(orient="records")
    if isinstance(tabla, list):
        return [fila for fila in tabla if isinstance(fila, dict) and fila.get("indice_proyectado") is not None]
    return []


def _nota_proyeccion_html(r: dict[str, Any]) -> str:
    if not r.get("icociv_final_es_proyectado"):
        return ""
    return (
        "<p><b>Nota:</b> El índice final ICOCIV fue obtenido mediante el módulo de proyección existente. "
        f"Modelo: {escape(str(r.get('modelo_proyeccion', '')))}. "
        f"Estado del horizonte: {escape(str(r.get('estado_horizonte', '')))}.</p>"
    )


def _trazabilidad_lineamientos(r: dict[str, Any]) -> str:
    lineas = []
    if r.get("factor_iccp") is not None:
        lineas.append(
            "R1 = (P - A) × [(I ICCP / I0 ICCP) - 1] = "
            f"{_fmt_moneda(r.get('base_ajustable'))} × [({_fmt_indice(r.get('i_iccp'))} / {_fmt_indice(r.get('i0_iccp'))}) - 1] = {_fmt_moneda(r.get('r1'))}"
        )
    else:
        lineas.append("R1 = No aplica")
    if r.get("factor_icociv") is not None:
        lineas.append(
            "R2 = [(P - A) + R1] × [(I ICOCIV / I0 ICOCIV) - 1] = "
            f"({_fmt_moneda(r.get('base_ajustable'))} + {_fmt_moneda(r.get('r1'))}) × "
            f"[({_fmt_indice(r.get('i_icociv'))} / {_fmt_indice(r.get('i0_icociv'))}) - 1] = {_fmt_moneda(r.get('r2'))}"
        )
    else:
        lineas.append("R2 = No aplica")
    lineas.append(f"R = R1 + R2 = {_fmt_moneda(r.get('r1'))} + {_fmt_moneda(r.get('r2'))} = {_fmt_moneda(r.get('r_total'))}")
    lineas.append(f"Valor insumo ajustado = (P - A) + R = {_fmt_moneda(r.get('base_ajustable'))} + {_fmt_moneda(r.get('r_total'))} = {_fmt_moneda(r.get('valor_actualizado'))}")
    return "\n".join(lineas)


def _dataframe_desde_registros(registros: list[dict[str, Any]], columnas: list[tuple[str, str]]) -> pd.DataFrame:
    titulos = [titulo for _, titulo in columnas]
    filas = [{titulo: registro.get(clave, "") for clave, titulo in columnas} for registro in registros]
    return pd.DataFrame(filas, columns=titulos)


def _activar_recalculo_excel(wb: Workbook) -> None:
    calculo = getattr(wb, "calculation", None) or getattr(wb, "calculation_properties", None)
    if calculo is not None:
        calculo.fullCalcOnLoad = True
        calculo.forceFullCalc = True


def _escribir_titulo_seccion_excel(ws, fila: int, titulo: str, columnas: int = 11) -> int:
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=columnas)
    celda = ws.cell(row=fila, column=1, value=titulo)
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="334155")
    celda.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[fila].height = 22
    return fila + 1


def _escribir_pares_excel(ws, fila: int, titulo: str, pares: list[tuple[str, Any]]) -> int:
    fila = _escribir_titulo_seccion_excel(ws, fila, titulo)
    borde = _borde_excel()
    for etiqueta, valor in pares:
        ws.cell(row=fila, column=1, value=etiqueta)
        ws.cell(row=fila, column=2, value=valor)
        for columna in (1, 2):
            celda = ws.cell(row=fila, column=columna)
            celda.border = borde
            celda.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=fila, column=1).font = Font(bold=True)
        fila += 1
    return fila + 1


def _escribir_tabla_excel(ws, fila: int, titulo: str, encabezados: list[str], filas: list[list[Any]]) -> int:
    fila = _escribir_titulo_seccion_excel(ws, fila, titulo, max(11, len(encabezados)))
    for columna, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila, column=columna, value=encabezado)
        _estilo_encabezado_tabla(celda)
    ws.row_dimensions[fila].height = 42
    for valores in filas:
        fila += 1
        for columna, valor in enumerate(valores, start=1):
            celda = ws.cell(row=fila, column=columna, value=valor)
            _estilo_celda_tabla(celda)
            if isinstance(valor, (int, float)):
                celda.number_format = FORMATO_INDICE_EXCEL
    return fila + 1


def _estilo_encabezado_tabla(celda) -> None:
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="475569")
    celda.border = _borde_excel()
    celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _estilo_celda_tabla(celda) -> None:
    celda.border = _borde_excel()
    celda.alignment = Alignment(vertical="center", wrap_text=True)


def _borde_excel() -> Border:
    lado = Side(style="thin", color="CBD5E1")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _ajustar_hoja_excel(ws) -> None:
    anchos = {
        "A": 26,
        "B": 22,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 12,
        "G": 18,
        "H": 18,
        "I": 12,
        "J": 14,
        "K": 18,
    }
    for indice in range(1, ws.max_column + 1):
        letra = get_column_letter(indice)
        ws.column_dimensions[letra].width = anchos.get(letra, 18)


def _numero_excel(valor: Any) -> float | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    if not texto or texto.lower().startswith("no aplica"):
        return None
    texto = texto.replace("$", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return None


def _errores_exportacion_excel(calculos: list[dict[str, Any]]) -> list[str]:
    errores: list[str] = []
    numericos = {
        "precio_base": "precio base",
        "r1": "R1",
        "r2": "R2",
        "r_total": "R",
    }
    for indice, calculo in enumerate(calculos, start=1):
        if not _insumo_contractual(calculo):
            errores.append(f"Cálculo {indice}: falta insumo contractual.")
        if not calculo.get("unidad") or calculo.get("unidad") == "Sin selección":
            errores.append(f"Cálculo {indice}: falta unidad válida.")
        if not calculo.get("fecha_inicial") or not calculo.get("fecha_final"):
            errores.append(f"Cálculo {indice}: faltan fechas para encabezados.")
        for clave, nombre in numericos.items():
            if _numero_excel(calculo.get(clave)) is None:
                errores.append(f"Cálculo {indice}: falta {nombre} numérico.")
    return errores


def _valor_detalle_exportacion(clave: str, valor: Any) -> Any:
    if clave == "diferencia_absoluta":
        return _fmt_moneda(valor)
    if clave == "advertencias_proyeccion" and isinstance(valor, list):
        return " | ".join(str(item) for item in valor)
    return valor


def _insumo_contractual(r: dict[str, Any]) -> str:
    return str(r.get("item") or r.get("insumo") or r.get("codigo_item") or "Sin insumo").strip()


def _equivalente_iccp(r: dict[str, Any]) -> str:
    serie = str(r.get("serie_iccp") or r.get("nombre_serie_iccp") or "").strip()
    if r.get("tipo_serie_iccp") == "canasta_general":
        return f"Canasta general > {serie}" if serie else "Canasta general"
    return serie or str(r.get("ruta_iccp") or "").strip()


def _equivalente_icociv(r: dict[str, Any]) -> str:
    for item in reversed(r.get("ruta_icociv_estructurada") or []):
        if isinstance(item, dict) and item.get("nivel") != "Tabla ICOCIV" and item.get("valor"):
            return _normalizar_nombre_icociv(item["valor"])
    ruta = str(r.get("ruta_icociv") or "")
    if ":" in ruta:
        ruta = ruta.split(":", 1)[1]
    partes = [parte.strip() for parte in ruta.split(">") if parte.strip()]
    return _normalizar_nombre_icociv(partes[-1]) if partes else ""


def _normalizar_nombre_icociv(texto: Any) -> str:
    return str(texto).replace("TEBERÍA", "TUBERÍA").replace("Tebería", "Tubería")
