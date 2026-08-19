"""Módulo ICOCIV: tabla acumulativa de proyecciones y calculadora de actualización.

Reutiliza el estilo, formato y exportación del módulo de empalme, pero trabaja
únicamente con índices ICOCIV (sin ICCP).
"""

from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any
from uuid import uuid4

import pandas as pd
from openpyxl import Workbook
from PySide6.QtWidgets import (
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTableView,
    QTextBrowser,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from app_icociv.interfaz.widgets.modelo_tabla import ModeloTablaPandas
from app_icociv.interfaz.widgets.controles import ComboBoxSinRueda
from app_icociv.interfaz.widgets.empalme_iccp_icociv import (
    FORMATO_INDICE_EXCEL,
    FORMATO_MONEDA_EXCEL,
    UNIDADES_INGENIERIA,
    _activar_recalculo_excel,
    _ajustar_hoja_excel,
    _ayuda,
    _escribir_pares_excel,
    _escribir_tabla_excel,
    _etiqueta_ayuda,
    _fmt_factor,
    _fmt_indice,
    _fmt_moneda,
    _fmt_porcentaje,
    _numero_excel,
    _periodo_visible,
    _spin,
)
from app_icociv.servicios.actualizacion_icociv import calcular_actualizacion_icociv


HOJA_RESUMEN = "Resumen"
HOJA_PROYECCIONES = "Proyecciones ICOCIV"
HOJA_VALORES = "Valores proyectados"
HOJA_TRAZABILIDAD = "Trazabilidad"
HOJA_METODOLOGIA = "Metodología y fórmulas"

# Columnas visibles de la tabla acumulativa de proyecciones.
# Se prioriza la ventana de proyección (periodo proyectado), no solo el horizonte.
COLUMNAS_PROYECCIONES = [
    ("numero", "N°"),
    ("item", "Ítem / insumo"),
    ("item_ruta", "Ruta ICOCIV resumida"),
    ("tabla_icociv", "Tabla ICOCIV"),
    ("fecha_final_serie", "Fecha final de serie histórica"),
    ("fecha_inicial_proyeccion", "Fecha inicial de proyección"),
    ("fecha_final_proyeccion", "Fecha final de proyección"),
    ("horizonte_solicitado", "Horizonte solicitado"),
    ("maximo_recomendado", "Alcance máximo de proyección"),
    ("modelo", "Modelo seleccionado"),
    ("indice_base", "Índice base"),
    ("indice_proyectado", "Índice proyectado"),
    ("estado", "Estado del horizonte"),
    # P0-C / C2: la columna entregaba la pareja [inferior, superior].
    ("ic95", "Intervalo de predicción"),
    ("mae", "MAE"),
    ("rmse", "RMSE"),
    ("mape", "MAPE"),
    ("smape", "sMAPE"),
    ("mase", "MASE"),
    ("advertencia", "Advertencia"),
    ("fecha_calculo", "Fecha de cálculo"),
]

# Columnas visibles de la tabla acumulativa de valores proyectados (calculadora).
COLUMNAS_VALORES = [
    ("numero", "N°"),
    ("item", "Ítem / insumo"),
    ("unidad", "Unidad"),
    ("cantidad", "Cantidad"),
    ("precio_base", "P"),
    ("anticipo", "A"),
    ("base_ajustable", "Base ajustable"),
    ("ruta_icociv", "Ruta ICOCIV"),
    ("i0_icociv", "I0 ICOCIV"),
    ("i_icociv", "I ICOCIV proyectado"),
    ("factor", "Factor"),
    ("r_total", "R"),
    ("valor_proyectado", "Valor proyectado"),
    ("diferencia", "Diferencia %"),
    ("modelo", "Modelo usado"),
    ("horizonte", "Horizonte"),
    ("fecha_calculo", "Fecha de cálculo"),
    ("observacion", "Observación técnica"),
]


class WidgetProyeccionesICOCIV(QWidget):
    """Calculadora ICOCIV y tablas acumulativas de la sesión."""

    def __init__(self) -> None:
        super().__init__()
        self.contexto: dict[str, Any] | None = None
        self.proyecciones: list[dict[str, Any]] = []
        self.valores: list[dict[str, Any]] = []
        self._item_editado = False  # el usuario escribió un nombre propio
        self.modelo_proyecciones = ModeloTablaPandas()
        self.modelo_valores = ModeloTablaPandas()
        self._crear_interfaz()
        self._actualizar_estado_calculadora()

    # ------------------------------------------------------------------ API
    def montar_controles_seleccion(self, grupo_seleccion, grupo_parametros, boton_ejecutar) -> None:
        """Recibe el selector jerárquico, los parámetros y el botón de ejecución
        creados por la ventana principal y los ubica al inicio del módulo."""
        encabezado = QLabel("1. Selección jerárquica ICOCIV y parámetros de proyección")
        encabezado.setObjectName("titulo_dashboard")
        ayuda = QLabel(
            "Seleccione la ruta ICOCIV, el horizonte y ejecute la proyección desde este módulo."
        )
        ayuda.setWordWrap(True)
        ayuda.setObjectName("ruta_dashboard")
        self._layout_seleccion.addWidget(encabezado)
        self._layout_seleccion.addWidget(ayuda)
        self._layout_seleccion.addWidget(grupo_seleccion)
        self._layout_seleccion.addWidget(grupo_parametros)
        fila = QHBoxLayout()
        fila.addStretch()
        fila.addWidget(boton_ejecutar)
        self._layout_seleccion.addLayout(fila)

    def actualizar_contexto(self, resultado_ui: dict[str, Any] | None, registrar: bool = True) -> None:
        """Recibe el resultado actual; acumula una fila si ``registrar`` y es válida."""
        contexto = _extraer_contexto(resultado_ui) if resultado_ui else None
        self.contexto = contexto
        self._actualizar_estado_calculadora()
        if registrar and contexto and contexto.get("proyeccion_generada"):
            self._registrar_proyeccion(contexto)

    # ---------------------------------------------------------------- UI
    def _crear_interfaz(self) -> None:
        layout_principal = QVBoxLayout(self)
        layout_principal.setContentsMargins(0, 0, 0, 0)
        contenido = QWidget()
        layout = QVBoxLayout(contenido)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(12)

        titulo = QLabel("Proyecciones ICOCIV y actualización de valores")
        titulo.setObjectName("titulo_dashboard")
        subtitulo = QLabel(
            "Cada proyección válida se acumula abajo. La calculadora usa el índice base "
            "(último observado) y el índice proyectado por la aplicación; no usa ICCP."
        )
        subtitulo.setWordWrap(True)
        layout.addWidget(titulo)
        layout.addWidget(subtitulo)

        # Contenedor donde la ventana principal monta el selector jerárquico,
        # los parámetros y el botón de ejecución (sección 1 del módulo).
        self.contenedor_seleccion = QWidget()
        self._layout_seleccion = QVBoxLayout(self.contenedor_seleccion)
        self._layout_seleccion.setContentsMargins(0, 0, 0, 0)
        self._layout_seleccion.setSpacing(10)
        layout.addWidget(self.contenedor_seleccion)

        self.lbl_contexto = QLabel()
        self.lbl_contexto.setWordWrap(True)
        self.lbl_contexto.setObjectName("ruta_dashboard")
        layout.addWidget(self.lbl_contexto)

        layout.addWidget(self._grupo_calculadora())
        layout.addWidget(self._grupo_resultado())
        layout.addWidget(self._grupo_tablas(), 1)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(contenido)
        layout_principal.addWidget(scroll)

    def _grupo_calculadora(self) -> QGroupBox:
        grupo = QGroupBox("Calculadora de actualización con proyección ICOCIV")
        grid = QGridLayout(grupo)
        self.item = QLineEdit()
        self.item.textEdited.connect(self._marcar_item_editado)
        self.unidad = ComboBoxSinRueda()
        self.unidad.addItem("Sin selección", None)
        for unidad in UNIDADES_INGENIERIA:
            self.unidad.addItem(unidad, unidad)
        self.cantidad = _spin(0, 1_000_000_000, 4)
        self.precio_base = _spin(0, 1_000_000_000_000_000, 2)
        self.anticipo = _spin(0, 1_000_000_000_000_000, 2)
        self.precio_base.lineEdit().setPlaceholderText("Ej.: 423467.54")
        self.anticipo.lineEdit().setPlaceholderText("Ej.: 0")
        self.observacion = QTextEdit()
        self.observacion.setMaximumHeight(64)
        campos = [
            ("Nombre del ítem o insumo", self.item, "Insumo, actividad, APU o componente contractual."),
            ("Unidad", self.unidad, "Unidad de medida del ítem."),
            ("Cantidad", self.cantidad, "Cantidad del insumo o ítem. Es opcional."),
            ("Precio o valor base P", self.precio_base, "Valor actual que se desea proyectar."),
            ("Anticipo amortizado A", self.anticipo, "Anticipo ya amortizado; si no aplica, deje en cero."),
            ("Observación técnica", self.observacion, "Justificación o criterio técnico del cálculo."),
        ]
        for fila, (etiqueta, widget, ayuda) in enumerate(campos):
            grid.addWidget(_etiqueta_ayuda(etiqueta, ayuda), fila, 0)
            grid.addWidget(_ayuda(widget, ayuda), fila, 1)
        grid.setColumnStretch(1, 1)

        acciones = QHBoxLayout()
        acciones.addStretch()
        self.boton_limpiar = QPushButton("Limpiar")
        self.boton_limpiar.clicked.connect(self._limpiar_calculadora)
        self.boton_calcular = QPushButton("Calcular y agregar a la tabla")
        self.boton_calcular.clicked.connect(self._calcular)
        acciones.addWidget(self.boton_limpiar)
        acciones.addWidget(self.boton_calcular)
        grid.addLayout(acciones, len(campos), 0, 1, 2)
        return grupo

    def _grupo_resultado(self) -> QGroupBox:
        grupo = QGroupBox("Resultado de actualización ICOCIV proyectada")
        layout = QVBoxLayout(grupo)
        self.resultado = QTextBrowser()
        self.resultado.setMinimumHeight(180)
        self.resultado.setHtml(
            "<p>Ejecute una proyección y calcule para ver P, A, base, I0, I, factor, R y el valor proyectado.</p>"
        )
        layout.addWidget(self.resultado)
        return grupo

    def _grupo_tablas(self) -> QGroupBox:
        grupo = QGroupBox("Tablas acumulativas de la sesión")
        layout = QVBoxLayout(grupo)

        layout.addWidget(QLabel("1. Proyecciones ICOCIV realizadas"))
        self.tabla_proyecciones = QTableView()
        self.tabla_proyecciones.setModel(self.modelo_proyecciones)
        self.tabla_proyecciones.setAlternatingRowColors(True)
        self.tabla_proyecciones.setMinimumHeight(150)
        self.tabla_proyecciones.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.tabla_proyecciones, 1)
        acc_proy = QHBoxLayout()
        boton_detalle_proy = QPushButton("Ver detalle")
        boton_detalle_proy.clicked.connect(self._detalle_proyeccion)
        boton_eliminar_proy = QPushButton("Eliminar")
        boton_eliminar_proy.clicked.connect(self._eliminar_proyeccion)
        for boton in (boton_detalle_proy, boton_eliminar_proy):
            acc_proy.addWidget(boton)
        acc_proy.addStretch()
        layout.addLayout(acc_proy)

        layout.addWidget(QLabel("2. Valores proyectados ICOCIV (calculadora)"))
        self.tabla_valores = QTableView()
        self.tabla_valores.setModel(self.modelo_valores)
        self.tabla_valores.setAlternatingRowColors(True)
        self.tabla_valores.setMinimumHeight(150)
        self.tabla_valores.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.tabla_valores, 1)
        acc_val = QHBoxLayout()
        boton_detalle_val = QPushButton("Ver detalle")
        boton_detalle_val.clicked.connect(self._detalle_valor)
        boton_eliminar_val = QPushButton("Eliminar")
        boton_eliminar_val.clicked.connect(self._eliminar_valor)
        self.boton_exportar = QPushButton("Exportar Excel")
        self.boton_exportar.clicked.connect(self._exportar_excel)
        for boton in (boton_detalle_val, boton_eliminar_val, self.boton_exportar):
            acc_val.addWidget(boton)
        acc_val.addStretch()
        layout.addLayout(acc_val)
        return grupo

    # ------------------------------------------------------------ acciones
    def _marcar_item_editado(self, texto: str) -> None:
        # Si el usuario borra todo el campo, se vuelve a permitir la sugerencia.
        self._item_editado = bool(texto.strip())

    def _sugerir_nombre_item(self) -> None:
        """Rellena el nombre del ítem con la serie final, sin pisar ediciones."""
        serie_final = (self.contexto or {}).get("serie_final", "") if self.contexto else ""
        self.item.setPlaceholderText(serie_final or "Nombre del ítem o insumo")
        if serie_final and not self._item_editado:
            self.item.setText(serie_final)

    def _limpiar_calculadora(self) -> None:
        self._item_editado = False
        self.item.clear()
        self.unidad.setCurrentIndex(0)
        self.cantidad.setValue(0)
        self.precio_base.setValue(0)
        self.anticipo.setValue(0)
        self.observacion.clear()
        self._sugerir_nombre_item()

    def _calcular(self) -> None:
        contexto = self.contexto
        if not contexto or not contexto.get("proyeccion_generada"):
            QMessageBox.information(
                self,
                "Calculadora ICOCIV",
                "Ejecute primero una proyección válida antes de calcular la actualización.",
            )
            return
        unidad = self.unidad.currentData()
        nombre_item = self.item.text().strip() or contexto.get("serie_final", "")
        entrada = {
            "item": nombre_item,
            "unidad": str(unidad or ""),
            "cantidad": self.cantidad.value() or None,
            "precio_base": self.precio_base.value(),
            "anticipo_amortizado": self.anticipo.value(),
            "i0_icociv": contexto.get("indice_base"),
            "i_icociv": contexto.get("indice_proyectado"),
            "ruta_icociv": contexto.get("ruta_texto"),
            "modelo_proyeccion": contexto.get("modelo"),
            "horizonte": contexto.get("horizonte_solicitado"),
            "periodo_base": contexto.get("periodo_base"),
            "periodo_proyectado": contexto.get("periodo_proyectado"),
            "observacion_tecnica": self.observacion.toPlainText().strip(),
        }
        try:
            if not nombre_item:
                raise ValueError("Ingrese el nombre del ítem o insumo.")
            if not unidad:
                raise ValueError("Seleccione una unidad válida.")
            resultado = calcular_actualizacion_icociv(entrada)
        except ValueError as exc:
            QMessageBox.warning(self, "Validación ICOCIV", str(exc))
            return
        resultado["calculo_id"] = uuid4().hex[:8]
        self.valores.append(resultado)
        self._mostrar_resultado(resultado)
        self._refrescar_tabla_valores()

    def _mostrar_resultado(self, r: dict[str, Any]) -> None:
        filas = [
            ("Precio base P", _fmt_moneda(r.get("precio_base"))),
            ("Anticipo A", _fmt_moneda(r.get("anticipo_amortizado"))),
            ("Base ajustable P - A", _fmt_moneda(r.get("base_ajustable"))),
            ("I0 ICOCIV", _fmt_indice(r.get("i0_icociv"))),
            ("I ICOCIV proyectado", _fmt_indice(r.get("i_icociv"))),
            ("Factor ICOCIV", _fmt_factor(r.get("factor_icociv"))),
            ("Valor del ajuste R", _fmt_moneda(r.get("r_total"))),
            ("Valor proyectado", _fmt_moneda(r.get("valor_proyectado"))),
            ("Diferencia porcentual", _fmt_porcentaje(r.get("diferencia_porcentual"))),
        ]
        filas_html = "".join(
            f"<tr><td style='padding:2px 10px 2px 0'><b>{escape(k)}</b></td>"
            f"<td style='padding:2px 0'>{escape(str(v))}</td></tr>"
            for k, v in filas
        )
        traza = escape(str(r.get("trazabilidad_formula", ""))).replace("\n", "<br>")
        self.resultado.setHtml(
            f"<h3>Actualización ICOCIV proyectada</h3><table>{filas_html}</table>"
            f"<h4>Desarrollo matemático</h4><p style='font-family:monospace'>{traza}</p>"
        )

    def _detalle_proyeccion(self) -> None:
        indice = _seleccion(self.tabla_proyecciones)
        if indice is not None and 0 <= indice < len(self.proyecciones):
            r = self.proyecciones[indice]
            QMessageBox.information(self, "Detalle de la proyección", _detalle_texto(r, COLUMNAS_PROYECCIONES))

    def _detalle_valor(self) -> None:
        indice = _seleccion(self.tabla_valores)
        if indice is not None and 0 <= indice < len(self.valores):
            self._mostrar_resultado(self.valores[indice])

    def _eliminar_proyeccion(self) -> None:
        indice = _seleccion(self.tabla_proyecciones)
        if indice is not None and 0 <= indice < len(self.proyecciones):
            del self.proyecciones[indice]
            self._refrescar_tabla_proyecciones()

    def _eliminar_valor(self) -> None:
        indice = _seleccion(self.tabla_valores)
        if indice is not None and 0 <= indice < len(self.valores):
            del self.valores[indice]
            self._refrescar_tabla_valores()

    def _exportar_excel(self) -> None:
        if not self.proyecciones and not self.valores:
            QMessageBox.information(self, "Exportación", "No hay proyecciones ni valores para exportar.")
            return
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Exportar proyecciones ICOCIV", "proyecciones_icociv.xlsx", "Excel (*.xlsx)"
        )
        if not ruta:
            return
        if not ruta.lower().endswith(".xlsx"):
            ruta = f"{ruta}.xlsx"
        try:
            self._generar_excel(ruta)
        except Exception as exc:  # pragma: no cover - mensaje UI
            QMessageBox.critical(self, "Exportación", f"No fue posible generar el archivo:\n{exc}")
            return
        QMessageBox.information(self, "Exportación", f"Archivo generado: {Path(ruta).name}")

    # ------------------------------------------------------------ internos
    def _registrar_proyeccion(self, contexto: dict[str, Any]) -> None:
        fila = _fila_proyeccion(contexto)
        # ponytail: dedupe contra la última fila por (ruta, periodo, horizonte)
        if self.proyecciones:
            ultima = self.proyecciones[-1]
            clave = ("item_ruta", "periodo_proyectado", "horizonte_solicitado")
            if all(ultima.get(k) == fila.get(k) for k in clave):
                self.proyecciones[-1] = fila
                self._refrescar_tabla_proyecciones()
                return
        self.proyecciones.append(fila)
        self._refrescar_tabla_proyecciones()

    def _refrescar_tabla_proyecciones(self) -> None:
        for numero, fila in enumerate(self.proyecciones, start=1):
            fila["numero"] = numero
        self.modelo_proyecciones.establecer_dataframe(_dataframe(self.proyecciones, COLUMNAS_PROYECCIONES))

    def _refrescar_tabla_valores(self) -> None:
        for numero, fila in enumerate(self.valores, start=1):
            fila["numero_fila"] = numero
        self.modelo_valores.establecer_dataframe(_dataframe_valores(self.valores))

    def _actualizar_estado_calculadora(self) -> None:
        contexto = self.contexto
        valido = bool(contexto and contexto.get("proyeccion_generada"))
        self.boton_calcular.setEnabled(valido)
        if not contexto:
            self.lbl_contexto.setText("Sin proyección activa. Ejecute una proyección para habilitar la calculadora.")
        elif not valido:
            self.lbl_contexto.setText(
                f"Proyección no generada para {contexto.get('ruta_texto', 'la ruta actual')}. "
                "La calculadora permanece deshabilitada."
            )
        else:
            self.lbl_contexto.setText(
                f"Proyección activa: {contexto.get('ruta_texto', '')} · "
                f"I0={_fmt_indice(contexto.get('indice_base'))} ({_periodo_visible(contexto.get('periodo_base'))}) · "
                f"I proyectado={_fmt_indice(contexto.get('indice_proyectado'))} ({_periodo_visible(contexto.get('periodo_proyectado'))}) · "
                f"Modelo: {contexto.get('modelo', 'No aplica')}"
            )
        self._sugerir_nombre_item()

    def _generar_excel(self, ruta: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = HOJA_RESUMEN
        _activar_recalculo_excel(wb)
        contexto = self.contexto or {}
        _escribir_pares_excel(
            ws,
            1,
            "Resumen de la sesión ICOCIV",
            [
                ("Fecha de exportación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("Proyecciones acumuladas", len(self.proyecciones)),
                ("Valores proyectados acumulados", len(self.valores)),
                ("Tabla ICOCIV activa", contexto.get("tabla_icociv", "No aplica")),
                ("Ruta jerárquica seleccionada", contexto.get("ruta_texto", "No aplica")),
                ("Serie final seleccionada", contexto.get("serie_final", "No aplica")),
                ("Fecha final de serie histórica", _periodo_iso(contexto.get("fecha_final_serie")) if contexto.get("fecha_final_serie") else "No aplica"),
                ("Fecha inicial de proyección", contexto.get("fecha_inicial_proyeccion", "No aplica")),
                ("Fecha final de proyección", contexto.get("fecha_final_proyeccion", "No aplica")),
                ("Modelo activo", contexto.get("modelo", "No aplica")),
            ],
        )

        ws_proy = wb.create_sheet(HOJA_PROYECCIONES)
        _escribir_tabla_excel(
            ws_proy,
            1,
            "Proyecciones ICOCIV de la sesión",
            [titulo for _, titulo in COLUMNAS_PROYECCIONES],
            [[fila.get(clave, "") for clave, _ in COLUMNAS_PROYECCIONES] for fila in self.proyecciones],
        )

        ws_val = wb.create_sheet(HOJA_VALORES)
        self._escribir_valores_excel(ws_val)

        ws_traza = wb.create_sheet(HOJA_TRAZABILIDAD)
        self._escribir_trazabilidad_excel(ws_traza)

        ws_met = wb.create_sheet(HOJA_METODOLOGIA)
        _escribir_pares_excel(
            ws_met,
            1,
            "Metodología ICOCIV",
            [
                ("Base ajustable", "Base = P - A"),
                ("Factor ICOCIV", "Factor = I / I0"),
                ("Valor del ajuste R", "R = (P - A) × [(I / I0) - 1]"),
                ("Valor proyectado", "Valor proyectado = (P - A) + R"),
                ("Índices", "I0: índice ICOCIV del último periodo observado. I: índice ICOCIV proyectado por la aplicación."),
                ("Alcance", "Cálculo de un solo tramo ICOCIV; no interviene ICCP."),
            ],
        )

        for hoja in wb.worksheets:
            _ajustar_hoja_excel(hoja)
        Path(ruta).parent.mkdir(parents=True, exist_ok=True)
        wb.save(ruta)

    def _escribir_valores_excel(self, ws) -> None:
        encabezados = [titulo for _, titulo in COLUMNAS_VALORES]
        from app_icociv.interfaz.widgets.empalme_iccp_icociv import (
            _escribir_titulo_seccion_excel,
            _estilo_celda_tabla,
            _estilo_encabezado_tabla,
        )

        fila = _escribir_titulo_seccion_excel(ws, 1, "Valores proyectados ICOCIV", max(11, len(encabezados)))
        for columna, titulo in enumerate(encabezados, start=1):
            _estilo_encabezado_tabla(ws.cell(row=fila, column=columna, value=titulo))
        ws.row_dimensions[fila].height = 40
        base_datos = fila + 1
        for indice, r in enumerate(self.valores):
            fila += 1
            f = fila
            valores = [
                indice + 1,
                r.get("item", ""),
                r.get("unidad", ""),
                _numero_excel(r.get("cantidad")),
                _numero_excel(r.get("precio_base")),
                _numero_excel(r.get("anticipo_amortizado")),
                f"=E{f}-F{f}",  # base = P - A
                r.get("ruta_icociv", ""),
                _numero_excel(r.get("i0_icociv")),
                _numero_excel(r.get("i_icociv")),
                f"=IF(I{f}=0,0,J{f}/I{f})",  # factor = I / I0
                f'=IF(I{f}=0,0,(E{f}-F{f})*((J{f}/I{f})-1))',  # R
                f"=(E{f}-F{f})+L{f}",  # valor proyectado
                f'=IF((E{f}-F{f})=0,0,(L{f}/(E{f}-F{f}))*100)',  # diferencia %
                r.get("modelo_proyeccion", ""),
                r.get("horizonte", ""),
                str(r.get("fecha_calculo", "")),
                r.get("observacion_tecnica", ""),
            ]
            for columna, valor in enumerate(valores, start=1):
                _estilo_celda_tabla(ws.cell(row=f, column=columna, value=valor))
            for columna in (5, 6, 7, 12, 13):
                ws.cell(row=f, column=columna).number_format = FORMATO_MONEDA_EXCEL
            for columna in (9, 10):
                ws.cell(row=f, column=columna).number_format = FORMATO_INDICE_EXCEL
        _ = base_datos

    def _escribir_trazabilidad_excel(self, ws) -> None:
        filas = []
        for indice, r in enumerate(self.valores, start=1):
            filas.append([indice, r.get("item", ""), str(r.get("trazabilidad_formula", ""))])
        _escribir_tabla_excel(ws, 1, "Trazabilidad de fórmulas", ["N°", "Ítem", "Desarrollo matemático"], filas)


# ------------------------------------------------------------- utilidades


def _extraer_contexto(resultado_ui: dict[str, Any]) -> dict[str, Any]:
    proyeccion = resultado_ui.get("proyeccion", {}) or {}
    solicitado = proyeccion.get("resultado_horizonte_solicitado", {}) or {}
    info = proyeccion.get("analisis_horizontes_completo") or proyeccion.get("horizonte_info") or {}
    if "alcance_maximo_proyeccion" not in info:
        info = proyeccion.get("horizonte_info", {}) or {}
    serie_df = resultado_ui.get("serie_df")
    indice_base = None
    periodo_base = ""
    periodo_inicial_serie = ""
    if isinstance(serie_df, pd.DataFrame) and not serie_df.empty and "Indice" in serie_df.columns:
        indice_base = float(serie_df["Indice"].iloc[-1])
        if "Periodo" in serie_df.columns:
            periodo_base = str(serie_df["Periodo"].iloc[-1])
            periodo_inicial_serie = str(serie_df["Periodo"].iloc[0])
    ruta = resultado_ui.get("ruta_jerarquica") or []
    ruta_sin_tabla = [item for item in ruta if item.get("nivel") != "Tabla ICOCIV"]
    metricas = (proyeccion.get("backtesting", {}) or {}).get("metricas", {}) or {}
    periodo_ini_proy, periodo_fin_proy = _ventana_proyeccion(proyeccion, periodo_base)
    return {
        "proyeccion_generada": bool(solicitado.get("proyeccion_generada")) and solicitado.get("indice_proyectado") is not None,
        "ruta": ruta,
        "ruta_texto": " · ".join(item.get("valor", "") for item in ruta_sin_tabla) or "Ruta no disponible",
        "serie_final": ruta_sin_tabla[-1]["valor"] if ruta_sin_tabla else "",
        "tabla_icociv": resultado_ui.get("fuente_visible", ""),
        "indice_base": indice_base,
        "periodo_base": periodo_base,
        "indice_proyectado": solicitado.get("indice_proyectado"),
        "periodo_proyectado": periodo_fin_proy or solicitado.get("periodo_proyectado") or proyeccion.get("periodo_proj", ""),
        "periodo_inicial_serie": periodo_inicial_serie,
        "fecha_final_serie": periodo_base,
        "fecha_inicial_proyeccion": periodo_ini_proy,
        "fecha_final_proyeccion": periodo_fin_proy,
        "horizonte_solicitado": solicitado.get("horizonte_solicitado"),
        "maximo_recomendado": info.get("alcance_maximo_proyeccion"),
        "estado": solicitado.get("estado"),
        "modelo": solicitado.get("modelo_aplicado") or proyeccion.get("model_name"),
        "metricas": metricas,
        "advertencia": solicitado.get("razon_principal") or "",
    }


def _ventana_proyeccion(proyeccion: dict[str, Any], periodo_base: str) -> tuple[str, str]:
    """Devuelve (primer periodo proyectado, último periodo proyectado) en ISO.

    Usa la tabla de proyecciones de la app cuando existe; si no, calcula desde
    el último periodo observado + horizonte.
    """
    proy_df = proyeccion.get("proyecciones")
    if isinstance(proy_df, pd.DataFrame) and not proy_df.empty and "periodo" in proy_df.columns:
        periodos = [str(p) for p in proy_df["periodo"].tolist() if str(p).strip()]
        if periodos:
            return _periodo_iso(periodos[0]), _periodo_iso(periodos[-1])
    solicitado = proyeccion.get("resultado_horizonte_solicitado", {}) or {}
    horizonte = solicitado.get("horizonte_solicitado") or proyeccion.get("horizonte_solicitado")
    if periodo_base and horizonte:
        try:
            anio, mes = (int(x) for x in str(periodo_base).replace("-", "_").split("_")[:2])
            inicio = _sumar_meses(anio, mes, 1)
            fin = _sumar_meses(anio, mes, int(horizonte))
            return inicio, fin
        except (TypeError, ValueError):
            return "", ""
    return "", ""


def _sumar_meses(anio: int, mes: int, delta: int) -> str:
    total = anio * 12 + (mes - 1) + delta
    return f"{total // 12:04d}-{total % 12 + 1:02d}"


def _periodo_iso(periodo: Any) -> str:
    texto = str(periodo or "").strip().replace("/", "_").replace("-", "_")
    partes = [p for p in texto.split("_") if p]
    if len(partes) >= 2 and partes[0].isdigit() and partes[1].isdigit():
        return f"{int(partes[0]):04d}-{int(partes[1]):02d}"
    return str(periodo or "")


_ESTADOS = {
    "proyeccion_tecnica": "Proyección técnica",
    "escenario": "Escenario",
    "no_admisible": "No admisible",
}


def _fila_proyeccion(contexto: dict[str, Any]) -> dict[str, Any]:
    metricas = contexto.get("metricas", {}) or {}
    maximo = contexto.get("maximo_recomendado")
    return {
        "calculo_id": uuid4().hex[:8],
        "item": contexto.get("serie_final", ""),
        "item_ruta": contexto.get("ruta_texto", ""),
        "tabla_icociv": contexto.get("tabla_icociv", ""),
        "serie_final": contexto.get("serie_final", ""),
        "fecha_final_serie": _periodo_iso(contexto.get("fecha_final_serie")),
        "fecha_inicial_proyeccion": contexto.get("fecha_inicial_proyeccion", ""),
        "fecha_final_proyeccion": contexto.get("fecha_final_proyeccion", ""),
        "periodo_proyectado": contexto.get("periodo_proyectado", ""),
        "indice_base": _fmt_indice(contexto.get("indice_base")),
        "indice_proyectado": _fmt_indice(contexto.get("indice_proyectado")),
        "periodo_proyectado": contexto.get("periodo_proyectado", ""),
        "horizonte_solicitado": _entero(contexto.get("horizonte_solicitado")),
        "maximo_recomendado": f"{int(maximo)} meses" if maximo else "No identificado",
        "estado": _ESTADOS.get(str(contexto.get("estado")), "No evaluado"),
        "modelo": str(contexto.get("modelo") or "No aplica"),
        # P0-C / C2: no se publica intervalo; se dice, no se deja "No aplica".
        "ic95": "No se publica en esta versión",
        "mae": _fmt_indice(metricas.get("mae")),
        "rmse": _fmt_indice(metricas.get("rmse")),
        "mape": _fmt_porcentaje(metricas.get("mape")),
        "smape": _fmt_porcentaje(metricas.get("smape")),
        "mase": _fmt_indice(metricas.get("mase")),
        "advertencia": str(contexto.get("advertencia") or ""),
        "fecha_calculo": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _dataframe_valores(valores: list[dict[str, Any]]) -> pd.DataFrame:
    filas = []
    for numero, r in enumerate(valores, start=1):
        filas.append(
            {
                "N°": numero,
                "Ítem / insumo": r.get("item", ""),
                "Unidad": r.get("unidad", ""),
                "Cantidad": r.get("cantidad", ""),
                "P": _fmt_moneda(r.get("precio_base")),
                "A": _fmt_moneda(r.get("anticipo_amortizado")),
                "Base ajustable": _fmt_moneda(r.get("base_ajustable")),
                "Ruta ICOCIV": r.get("ruta_icociv", ""),
                "I0 ICOCIV": _fmt_indice(r.get("i0_icociv")),
                "I ICOCIV proyectado": _fmt_indice(r.get("i_icociv")),
                "Factor": _fmt_factor(r.get("factor_icociv")),
                "R": _fmt_moneda(r.get("r_total")),
                "Valor proyectado": _fmt_moneda(r.get("valor_proyectado")),
                "Diferencia %": _fmt_porcentaje(r.get("diferencia_porcentual")),
                "Modelo usado": r.get("modelo_proyeccion", ""),
                "Horizonte": r.get("horizonte", ""),
                "Fecha de cálculo": r.get("fecha_calculo", ""),
                "Observación técnica": r.get("observacion_tecnica", ""),
            }
        )
    columnas = [titulo for _, titulo in COLUMNAS_VALORES]
    return pd.DataFrame(filas, columns=columnas)


def _dataframe(filas: list[dict[str, Any]], columnas: list[tuple[str, str]]) -> pd.DataFrame:
    titulos = [titulo for _, titulo in columnas]
    registros = [{titulo: fila.get(clave, "") for clave, titulo in columnas} for fila in filas]
    return pd.DataFrame(registros, columns=titulos)


def _detalle_texto(fila: dict[str, Any], columnas: list[tuple[str, str]]) -> str:
    return "\n".join(f"{titulo}: {fila.get(clave, '')}" for clave, titulo in columnas)


def _seleccion(tabla: QTableView) -> int | None:
    indices = tabla.selectionModel().selectedRows() if tabla.selectionModel() else []
    if indices:
        return indices[0].row()
    actual = tabla.currentIndex()
    return actual.row() if actual.isValid() else None


def _entero(valor: Any) -> Any:
    try:
        return int(valor)
    except (TypeError, ValueError):
        return "No disponible"
